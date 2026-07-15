import json
import logging
import re
import threading
from datetime import datetime, timedelta

from django.contrib.auth.decorators import login_required

logger = logging.getLogger(__name__)

from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import send_mail
from django.db.models import Q, Count, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import (
    CreateView, DetailView, ListView, TemplateView, UpdateView, View
)

from .forms import (ClienteForm, DocumentoForm, ExpedienteForm, NotaForm,
                       SolicitudConciliacionForm, WhatsAppMessageForm,
                       CalculoLaboralForm, SimulacionForm)
from .models import (Cliente, Documento, Expediente, Movimiento, Nota,
                      SolicitudConciliacion, WhatsAppMessage, TareaConciliacion,
                      CalculoLaboral, LegalConfig, Aviso,
                      SolicitudTransferencia, Notificacion)
from .signals import registrar_movimiento
from .whatsapp import (enviar_whatsapp, generar_deep_link, renderizar_plantilla,
                        MENSAJES_TEMPLATE)
from .laboral_calculator import calcular_desde_expediente, recalcular_calculo
from .demanda_generator import generar_demanda_word, generar_demanda_html, html_a_docx, PLANTILLAS_INFO
from .models import Machote
from .marcadores import get_marcadores, get_datos_faltantes, get_completitud_stats, reemplazar_marcadores


def _get_machotes_queryset():
    """Retorna Machotes activos ordenados: favoritos primero, luego por orden y nombre."""
    return Machote.objects.filter(activo=True).order_by('-favorito', 'orden', 'nombre')
from core.laboral.calculators import simular
from .conciliacion_automation import enviar_y_guardar as enviar_conciliacion


# ─── Mixins ────────────────────────────────────────────────────────────────

class AsesorOnlyMixin(UserPassesTestMixin):
    """Restringe acceso a usuarios con rol asesor."""
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.rol == 'asesor'


class AdminOrSuperOnlyMixin(UserPassesTestMixin):
    """Restringe acceso a administrativos y superadmin."""
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.rol in ['admin', 'superadmin']


class StaffRequiredMixin(UserPassesTestMixin):
    """Cualquier usuario autenticado del sistema."""
    def test_func(self):
        return self.request.user.is_authenticated


# ─── Helpers ───────────────────────────────────────────────────────────────

def get_expedientes_queryset(user):
    """Retorna queryset de expedientes según el rol del usuario."""
    if hasattr(user, 'profile') and user.profile.rol in ['admin', 'superadmin']:
        return Expediente.objects.select_related('cliente', 'asesor').all()
    return Expediente.objects.select_related('cliente', 'asesor').filter(asesor=user)


def get_clientes_queryset(user):
    """Retorna queryset de clientes según el rol."""
    if hasattr(user, 'profile') and user.profile.rol in ['admin', 'superadmin']:
        return Cliente.objects.all()
    return Cliente.objects.filter(expediente__asesor=user).distinct()


ESTADO_COLORS = {
    'nuevo': 'bg-blue-100 text-blue-700',
    'solicitud': 'bg-purple-100 text-purple-700',
    'citatorio': 'bg-indigo-100 text-indigo-700',
    'audiencia': 'bg-yellow-100 text-yellow-700',
    'no_notificado': 'bg-red-100 text-red-700',
    'reprogramacion': 'bg-orange-100 text-orange-700',
    'convenio': 'bg-green-100 text-green-700',
    'sin_conciliacion': 'bg-gray-100 text-gray-700',
    'demanda': 'bg-rose-100 text-rose-700',
    'cerrado': 'bg-slate-100 text-slate-700',
}


# ─── Redirect ──────────────────────────────────────────────────────────────

@login_required
def dashboard_redirect(request):
    if hasattr(request.user, 'profile'):
        rol = request.user.profile.rol
        if rol == 'asesor':
            return redirect('dashboard_asesor')
        elif rol == 'finanzas':
            return redirect('dashboard_financiero')
    return redirect('dashboard_admin')


# ─── Dashboards ────────────────────────────────────────────────────────────

class DashboardAsesorView(LoginRequiredMixin, TemplateView):
    template_name = 'expedientes/dashboard_asesor.html'

    def get(self, request, *args, **kwargs):
        if not hasattr(request.user, 'profile') or request.user.profile.rol != 'asesor':
            return redirect('dashboard_admin')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        qs = Expediente.objects.filter(asesor=user)

        context['total_casos'] = qs.count()
        context['casos_activos'] = qs.exclude(estado='cerrado').count()
        context['casos_cerrados'] = qs.filter(estado='cerrado').count()
        context['proximas_audiencias'] = qs.filter(
            fecha_audiencia__isnull=False
        ).exclude(estado='cerrado').order_by('fecha_audiencia')[:5]
        context['mis_casos'] = qs.order_by('-created_at')[:10]
        context['alertas'] = qs.filter(prioridad='alta', estado__in=['nuevo', 'solicitud', 'audiencia']).count()
        context['casos_pendientes'] = qs.filter(estado__in=['nuevo', 'solicitud']).count()

        # Proximas asesorias gratuitas
        context['proximas_asesorias'] = Cliente.objects.filter(
            asesoria_gratuita_agendada=True,
            fecha_asesoria_gratuita__isnull=False,
        ).order_by('fecha_asesoria_gratuita')[:10]

        # Avisos activos para todos (ordenados por prioridad: alta, media, baja)
        from django.db.models import Case, When, IntegerField, Value
        context['avisos'] = Aviso.objects.filter(activo=True).annotate(
            prioridad_num=Case(
                When(prioridad='alta', then=Value(0)),
                When(prioridad='media', then=Value(1)),
                When(prioridad='baja', then=Value(2)),
                output_field=IntegerField(),
            )
        ).order_by('prioridad_num', '-created_at')[:10]

        context['ESTADO_COLORS'] = ESTADO_COLORS
        return context


class DashboardAdminView(LoginRequiredMixin, StaffRequiredMixin, TemplateView):
    template_name = 'expedientes/dashboard_admin.html'

    def get(self, request, *args, **kwargs):
        if hasattr(request.user, 'profile') and request.user.profile.rol == 'asesor':
            return redirect('dashboard_asesor')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qs = Expediente.objects.all()

        context['total_casos'] = qs.count()
        context['casos_activos'] = qs.exclude(estado='cerrado').count()
        context['casos_cerrados'] = qs.filter(estado='cerrado').count()
        context['convenios_count'] = qs.filter(estado='convenio').count()
        context['demandas_count'] = qs.filter(estado='demanda').count()

        montos = qs.aggregate(
            total_reclamado=Sum('monto_reclamado'),
            total_convenio=Sum('monto_convenio')
        )
        context['total_reclamado'] = montos['total_reclamado'] or 0
        context['total_convenio'] = montos['total_convenio'] or 0

        # Productividad por asesor
        asesores = User.objects.filter(profile__rol='asesor').annotate(
            total_casos=Count('expediente'),
            casos_activos=Count('expediente', filter=~Q(expediente__estado='cerrado')),
            casos_cerrados=Count('expediente', filter=Q(expediente__estado='cerrado')),
            convenios=Count('expediente', filter=Q(expediente__estado='convenio')),
        )
        context['asesores'] = asesores

        context['proximas_audiencias'] = qs.filter(
            fecha_audiencia__isnull=False,
        ).exclude(estado='cerrado').order_by('fecha_audiencia')[:10]

        context['ultimos_movimientos'] = Movimiento.objects.select_related(
            'expediente', 'usuario'
        ).order_by('-created_at')[:15]

        # Transferencias pendientes
        context['transferencias_pendientes'] = SolicitudTransferencia.objects.filter(
            estado='pendiente'
        ).select_related(
            'expediente', 'expediente__cliente', 'solicitante'
        ).order_by('-created_at')[:10]

        context['casos_por_estado'] = {
            label: qs.filter(estado=key).count()
            for key, label in Expediente.ESTADO_CHOICES
        }

        context['proximas_asesorias'] = Cliente.objects.filter(
            asesoria_gratuita_agendada=True,
            fecha_asesoria_gratuita__isnull=False,
        ).order_by('fecha_asesoria_gratuita')[:10]

        # Avisos activos ordenados por prioridad (alta, media, baja)
        from django.db.models import Case, When, IntegerField, Value
        context['avisos'] = Aviso.objects.filter(activo=True).annotate(
            prioridad_num=Case(
                When(prioridad='alta', then=Value(0)),
                When(prioridad='media', then=Value(1)),
                When(prioridad='baja', then=Value(2)),
                output_field=IntegerField(),
            )
        ).order_by('prioridad_num', '-created_at')[:10]

        context['ESTADO_COLORS'] = ESTADO_COLORS

        return context


# ─── Búsqueda Global ───────────────────────────────────────────────────────

@login_required
def busqueda_global(request):
    q = request.GET.get('q', '').strip()
    resultados = None
    if q:
        expedientes = get_expedientes_queryset(request.user)
        resultados = expedientes.filter(
            Q(numero__icontains=q) |
            Q(cliente__nombre__icontains=q) |
            Q(cliente__empresa__icontains=q) |
            Q(cliente__telefono__icontains=q) |
            Q(cliente__curp__icontains=q) |
            Q(folio__icontains=q) |
            Q(estado__icontains=q)
        ).order_by('-created_at')[:20]

    return render(request, 'expedientes/busqueda_global.html', {
        'q': q,
        'resultados': resultados,
        'ESTADO_COLORS': ESTADO_COLORS,
    })


# ─── CRUD Expedientes ──────────────────────────────────────────────────────

class ExpedienteListView(LoginRequiredMixin, ListView):
    model = Expediente
    template_name = 'expedientes/expediente_list.html'
    paginate_by = 20

    def get_queryset(self):
        qs = get_expedientes_queryset(self.request.user)

        q = self.request.GET.get('q', '')
        estado = self.request.GET.get('estado', '')
        asesor_id = self.request.GET.get('asesor', '')
        fecha_desde = self.request.GET.get('fecha_desde', '')
        fecha_hasta = self.request.GET.get('fecha_hasta', '')
        prioridad = self.request.GET.get('prioridad', '')

        if q:
            qs = qs.filter(
                Q(numero__icontains=q) |
                Q(cliente__nombre__icontains=q) |
                Q(cliente__curp__icontains=q) |
                Q(cliente__empresa__icontains=q) |
                Q(folio__icontains=q)
            )
        if estado:
            qs = qs.filter(estado=estado)
        if asesor_id:
            qs = qs.filter(asesor_id=asesor_id)
        if prioridad:
            qs = qs.filter(prioridad=prioridad)
        if fecha_desde:
            qs = qs.filter(created_at__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(created_at__date__lte=fecha_hasta)

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estados'] = Expediente.ESTADO_CHOICES
        context['asesores_filtro'] = User.objects.filter(profile__rol='asesor')
        context['filtros'] = {k: v for k, v in self.request.GET.items() if v}
        context['ESTADO_COLORS'] = ESTADO_COLORS
        return context


class ExpedienteCreateView(LoginRequiredMixin, CreateView):
    model = Expediente
    form_class = ExpedienteForm
    template_name = 'expedientes/expediente_form.html'
    success_url = reverse_lazy('expediente_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        # Pre-seleccionar cliente si se pasa ?cliente=X
        cliente_id = self.request.GET.get('cliente')
        if cliente_id and not kwargs.get('data'):
            try:
                kwargs['initial'] = kwargs.get('initial', {})
                kwargs['initial']['cliente'] = int(cliente_id)
            except (ValueError, TypeError):
                pass
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente_id = self.request.GET.get('cliente')
        if cliente_id:
            try:
                context['cliente_preseleccionado'] = Cliente.objects.get(pk=cliente_id)
            except Cliente.DoesNotExist:
                pass
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        registrar_movimiento(
            expediente=self.object,
            usuario=self.request.user,
            accion='creacion',
            detalle=f'Expediente creado por {self.request.user.get_full_name() or self.request.user.username}'
        )
        messages.success(self.request, f'✅ Expediente {self.object.numero} creado. Ahora puedes programar la audiencia.')

        # Si el usuario eligió "Crear y Enviar a Conciliación", redirigir al flujo de conciliación
        if self.request.POST.get('action') == 'crear_y_conciliar':
            messages.info(self.request, '🚀 Expediente creado. Ahora puedes enviar la solicitud al portal de conciliación.')
            return redirect('enviar_conciliacion_automation', pk=self.object.pk)

        return response


class ExpedienteDetailView(LoginRequiredMixin, DetailView):
    model = Expediente
    template_name = 'expedientes/expediente_detail.html'

    def get_queryset(self):
        return get_expedientes_queryset(self.request.user).prefetch_related(
            'whatsapp_mensajes',
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['documentos'] = self.object.documentos.all()
        context['movimientos'] = self.object.movimientos.select_related('usuario').all()[:30]
        context['notas_lista'] = self.object.notas_lista.select_related('usuario').all()[:20]
        context['documento_form'] = DocumentoForm()
        context['nota_form'] = NotaForm()
        context['transiciones_posibles'] = [
            (key, dict(Expediente.ESTADO_CHOICES)[key])
            for key in Expediente.TRANSICIONES.get(self.object.estado, [])
        ]
        context['transferencias'] = self.object.solicitudes_transferencia.select_related(
            'solicitante', 'resuelto_por', 'asesor_asignado'
        ).order_by('-created_at')

        # Tareas de conciliación (últimas 5)
        context['tareas_conciliacion'] = self.object.tareas_conciliacion.select_related(
            'usuario'
        ).order_by('-created_at')[:5]
        context['ESTADO_COLORS'] = ESTADO_COLORS
        return context


class ExpedienteUpdateView(LoginRequiredMixin, UpdateView):
    model = Expediente
    form_class = ExpedienteForm
    template_name = 'expedientes/expediente_form.html'

    def get_queryset(self):
        return get_expedientes_queryset(self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_success_url(self):
        return reverse_lazy('expediente_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        old_estado = self.object.estado
        response = super().form_valid(form)

        if old_estado != self.object.estado:
            registrar_movimiento(
                expediente=self.object,
                usuario=self.request.user,
                accion='cambio_estado',
                detalle=f'Estado cambiado de "{dict(Expediente.ESTADO_CHOICES)[old_estado]}" a "{self.object.get_estado_display()}"'
            )
        else:
            registrar_movimiento(
                expediente=self.object,
                usuario=self.request.user,
                accion='actualizacion',
                detalle='Expediente actualizado'
            )
        return response


# ─── Cambio de estado rápido (HTMX) ────────────────────────────────────────

@login_required
def cambiar_estado(request, pk):
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado and nuevo_estado in Expediente.TRANSICIONES.get(expediente.estado, []):
            old_label = expediente.get_estado_display()
            expediente.estado = nuevo_estado
            expediente.save()
            registrar_movimiento(
                expediente=expediente,
                usuario=request.user,
                accion='cambio_estado',
                detalle=f'Estado cambiado de "{old_label}" a "{expediente.get_estado_display()}"'
            )
            return JsonResponse({
                'success': True,
                'estado': expediente.get_estado_display(),
                'estado_key': expediente.estado,
            })

    return JsonResponse({'success': False, 'error': 'Transición no permitida'}, status=400)


# ─── Resultado de Audiencia ────────────────────────────────────────────────

@login_required
def registrar_resultado_audiencia(request, pk):
    """Registra el resultado de una audiencia y cambia el estado según corresponda."""
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    if request.method == 'POST':
        resultado = request.POST.get('resultado')

        mapeo_resultados = {
            'no_notificado': 'no_notificado',
            'convenio': 'convenio',
            'sin_conciliacion': 'sin_conciliacion',
            'inasistencia': 'no_notificado',
            'reprogramada': 'reprogramacion',
        }

        nuevo_estado = mapeo_resultados.get(resultado)
        if nuevo_estado and nuevo_estado in Expediente.TRANSICIONES.get(expediente.estado, []):
            old_label = expediente.get_estado_display()
            expediente.estado = nuevo_estado
            expediente.resultado_audiencia = resultado

            # Si es convenio, registrar monto
            if resultado == 'convenio':
                monto = request.POST.get('monto_convenio')
                if monto:
                    expediente.monto_convenio = monto

            expediente.save()

            registrar_movimiento(
                expediente=expediente,
                usuario=request.user,
                accion='resultado_audiencia',
                detalle=f'Resultado de audiencia: {dict(Expediente.RESULTADO_AUDIENCIA_CHOICES).get(resultado, resultado)}. '
                        f'Estado cambiado a "{expediente.get_estado_display()}"'
            )

            return redirect('expediente_detail', pk=expediente.pk)

        return JsonResponse({'success': False, 'error': 'Resultado no válido para el estado actual'}, status=400)

    return redirect('expediente_detail', pk=expediente.pk)


# ─── Clientes ──────────────────────────────────────────────────────────────

class ClienteListView(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = 'expedientes/cliente_list.html'
    paginate_by = 20

    def get_queryset(self):
        qs = get_clientes_queryset(self.request.user)
        q = self.request.GET.get('q', '')
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) |
                Q(curp__icontains=q) |
                Q(empresa__icontains=q) |
                Q(telefono__icontains=q) |
                Q(email__icontains=q)
            )
        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtros'] = {k: v for k, v in self.request.GET.items() if v}
        return context


class ClienteCreateView(LoginRequiredMixin, CreateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'expedientes/cliente_form.html'

    def get_success_url(self):
        return reverse_lazy('expediente_create') + f'?cliente={self.object.pk}'


class ClienteUpdateView(LoginRequiredMixin, UpdateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'expedientes/cliente_form.html'
    success_url = reverse_lazy('cliente_list')


# ─── Documentos ────────────────────────────────────────────────────────────

@login_required
def subir_documento(request, pk):
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)
    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.expediente = expediente
            doc.subido_por = request.user
            doc.save()
            registrar_movimiento(
                expediente=expediente,
                usuario=request.user,
                accion='subida_documento',
                detalle=f'Documento subido: {doc.descripcion} ({doc.get_tipo_display()})'
            )
            messages.success(request, 'Documento subido correctamente.')
    return redirect('expediente_detail', pk=pk)


@login_required
def eliminar_documento(request, pk):
    doc = get_object_or_404(Documento, pk=pk)
    expediente = doc.expediente

    user = request.user
    if not (hasattr(user, 'profile') and user.profile.rol in ['admin', 'superadmin']):
        if expediente.asesor != user:
            return HttpResponse('No autorizado', status=403)

    doc.delete()
    registrar_movimiento(
        expediente=expediente,
        usuario=request.user,
        accion='actualizacion',
        detalle='Documento eliminado'
    )
    return redirect('expediente_detail', pk=expediente.pk)


# ─── Notas ─────────────────────────────────────────────────────────────────

@login_required
def agregar_nota(request, pk):
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)
    if request.method == 'POST':
        form = NotaForm(request.POST)
        if form.is_valid():
            nota = form.save(commit=False)
            nota.expediente = expediente
            nota.usuario = request.user
            nota.save()
            registrar_movimiento(
                expediente=expediente,
                usuario=request.user,
                accion='nota_agregada',
                detalle=f'Nota agregada: {nota.contenido[:100]}...'
            )
    return redirect('expediente_detail', pk=expediente.pk)


# ─── Calendario ────────────────────────────────────────────────────────────

@login_required
def calendario_audiencias(request):
    qs = get_expedientes_queryset(request.user)

    # Filtros
    asesor_id = request.GET.get('asesor', '')
    estado = request.GET.get('estado', '')
    vista = request.GET.get('vista', 'planner')

    if asesor_id:
        qs = qs.filter(asesor_id=asesor_id)
    if estado:
        qs = qs.filter(estado=estado)

    eventos = qs.filter(
        fecha_audiencia__isnull=False
    ).exclude(estado='cerrado')

    eventos_json = []
    for e in eventos.select_related('cliente', 'asesor'):
        eventos_json.append({
            'id': e.id,
            'title': f"{e.numero} - {e.cliente.nombre}",
            'start': e.fecha_audiencia.isoformat() if e.fecha_audiencia else None,
            'url': reverse_lazy('expediente_detail', kwargs={'pk': e.id}),
            'backgroundColor': {
                'audiencia': '#eab308',
                'no_notificado': '#ef4444',
                'reprogramacion': '#f97316',
                'convenio': '#22c55e',
            }.get(e.estado, '#3b82f6'),
            'borderColor': '#fff',
            'textColor': {
                'audiencia': '#854d0e',
                'no_notificado': '#fff',
                'reprogramacion': '#fff',
                'convenio': '#fff',
            }.get(e.estado, '#fff'),
        })

    # ── Planner context ──────────────────────────────────────────────
    hoy = timezone.now()
    # Soporte para navegación de semanas
    semana_offset = int(request.GET.get('semana', 0))
    inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
    dias_semana = []
    nombres_dias_es = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']  # 5-column design
    for i in range(5):  # Mon-Fri (user's design)
        dia = inicio_semana + timedelta(days=i)
        dias_semana.append({
            'fecha': dia,
            'numero': dia.day,
            'nombre': nombres_dias_es[i],
            'eventos': [],
        })

    # Agrupar eventos por día de la semana (Mon-Wed)
    for e in eventos.select_related('cliente', 'asesor').order_by('fecha_audiencia'):
        if e.fecha_audiencia:
            dia_evento = e.fecha_audiencia.weekday()
            if dia_evento < 5:  # Mon-Fri (0=Lun, 1=Mar, 2=Mie, 3=Jue, 4=Vie)
                dias_semana[dia_evento]['eventos'].append(e)

    mes_nombre = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                  'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    context_planner = {
        'dias_semana': dias_semana,
        'mes_actual': mes_nombre[inicio_semana.month - 1],
        'anio_actual': inicio_semana.year,
        'semana_actual': inicio_semana.isocalendar()[1],
        'semana_offset': semana_offset,
        'semana_anterior': semana_offset - 1,
        'semana_siguiente': semana_offset + 1,
        'inicio_semana': inicio_semana,
    }

    return render(request, 'expedientes/calendario.html', {
        'eventos_json': eventos_json,
        'eventos': eventos.order_by('fecha_audiencia')[:30],
        'asesores_filtro': User.objects.filter(profile__rol='asesor'),
        'estados_filtro': Expediente.ESTADO_CHOICES,
        'filtros': {'asesor': asesor_id, 'estado': estado, 'vista': vista},
        'vista_actual': vista,
        'ESTADO_COLORS': ESTADO_COLORS,
        **context_planner,
    })


# ─── Reportes ──────────────────────────────────────────────────────────────

@login_required
def exportar_excel(request):
    """Exporta expedientes a Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Expedientes'

        headers = ['N° Exp', 'Cliente', 'CURP', 'Empresa', 'Asesor', 'Estado', 'Monto Reclamado',
                    'Monto Convenio', 'Fecha Audiencia', 'Tipo Despido', 'Folio', 'Próxima Acción', 'Creado']
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        qs = get_expedientes_queryset(request.user)
        for i, exp in enumerate(qs, 2):
            ws.cell(row=i, column=1, value=exp.numero).border = thin_border
            ws.cell(row=i, column=2, value=exp.cliente.nombre).border = thin_border
            ws.cell(row=i, column=3, value=exp.cliente.curp).border = thin_border
            ws.cell(row=i, column=4, value=exp.cliente.empresa).border = thin_border
            ws.cell(row=i, column=5, value=exp.asesor.get_full_name() or exp.asesor.username).border = thin_border
            ws.cell(row=i, column=6, value=exp.get_estado_display()).border = thin_border
            ws.cell(row=i, column=7, value=float(exp.monto_reclamado or 0)).border = thin_border
            ws.cell(row=i, column=8, value=float(exp.monto_convenio or 0)).border = thin_border
            ws.cell(row=i, column=9, value=exp.fecha_audiencia.strftime('%d/%m/%Y %H:%M') if exp.fecha_audiencia else '').border = thin_border
            ws.cell(row=i, column=10, value=exp.get_tipo_despido_display() if exp.tipo_despido else '').border = thin_border
            ws.cell(row=i, column=11, value=exp.folio or '').border = thin_border
            ws.cell(row=i, column=12, value=exp.proxima_accion.strftime('%d/%m/%Y') if exp.proxima_accion else '').border = thin_border
            ws.cell(row=i, column=13, value=exp.created_at.strftime('%d/%m/%Y')).border = thin_border

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 22

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=expedientes_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    except ImportError:
        return HttpResponse('openpyxl no está instalado. Ejecute: pip install openpyxl', status=500)


@login_required
def generar_pdf_expediente(request, pk):
    """Genera PDF de un expediente."""
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)
    try:
        from weasyprint import HTML
        html_string = render_to_string('expedientes/pdf_expediente.html', {
            'expediente': expediente,
            'movimientos': expediente.movimientos.select_related('usuario').all()[:20],
            'documentos': expediente.documentos.all(),
            'hoy': timezone.now(),
        })
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{expediente.numero}.pdf"'
        HTML(string=html_string).write_pdf(response)
        return response
    except ImportError:
        return HttpResponse('WeasyPrint no está instalado. Ejecute: pip install weasyprint', status=500)


# ─── WhatsApp ────────────────────────────────────────────────────────────

@login_required
@require_POST
def toggle_whatsapp_auto(request, pk):
    """
    Activa/desactiva las notificaciones automáticas de WhatsApp para un expediente.
    Usa el estandar de Django para checkboxes: presente en POST = checked (True).
    """
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    expediente.notificar_whatsapp_auto = 'notificar_whatsapp_auto' in request.POST

    # Guardar solo este campo (evita validacion de transiciones de estado)
    expediente.save(update_fields=['notificar_whatsapp_auto'])

    return JsonResponse({
        'success': True,
        'notificar_whatsapp_auto': expediente.notificar_whatsapp_auto,
    })


@login_required
def whatsapp_enviar(request, pk):
    """Envía un mensaje de WhatsApp desde el expediente."""
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    if request.method == 'POST':
        form = WhatsAppMessageForm(request.POST)
        if form.is_valid():
            msg = form.save(commit=False)
            msg.expediente = expediente
            msg.enviado_por = request.user

            destino = msg.destino or expediente.cliente.whatsapp or expediente.cliente.telefono
            msg.destino = destino

            # Enviar el mensaje
            resultado = enviar_whatsapp(destino, msg.mensaje, via=msg.via)

            msg.link_generado = resultado.get('link', '')
            msg.estado = 'enviado' if resultado['success'] else 'fallido'
            if not resultado['success']:
                msg.error_log = resultado.get('detail', '')
            msg.save()

            registrar_movimiento(
                expediente=expediente,
                usuario=request.user,
                accion='actualizacion',
                detalle=f'WhatsApp enviado: {msg.get_tipo_display()} a {msg.destino}'
            )

            messages.success(request, f'✅ Mensaje de WhatsApp listo. Abre el enlace: {msg.link_generado}')
            return redirect('expediente_detail', pk=expediente.pk)
        else:
            messages.error(request, 'Corrige los errores del formulario.')
    else:
        # Pre-llenar con datos del cliente
        destino = expediente.cliente.whatsapp or expediente.cliente.telefono
        form = WhatsAppMessageForm(initial={
            'destino': destino,
            'mensaje': renderizar_plantilla('seguimiento',
                cliente=expediente.cliente.nombre,
                asesor=request.user.get_full_name() or request.user.username,
                fecha=''),
        })

    return render(request, 'expedientes/whatsapp_form.html', {
        'form': form,
        'expediente': expediente,
        'plantillas': MENSAJES_TEMPLATE,
        'ESTADO_COLORS': ESTADO_COLORS,
    })


@login_required
def whatsapp_historial(request, pk):
    """Retorna el historial de WhatsApp como HTML parcial (HTMX)."""
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)
    mensajes = expediente.whatsapp_mensajes.all()[:20]

    return render(request, 'expedientes/whatsapp_historial.html', {
        'mensajes': mensajes,
        'expediente': expediente,
    })


@login_required
def whatsapp_plantilla(request, tipo):
    """Retorna el texto de una plantilla de mensaje en JSON."""
    plantilla = renderizar_plantilla(tipo,
        cliente='[Nombre del cliente]',
        asesor='[Nombre del asesor]',
        fecha='[Fecha]',
    )
    return JsonResponse({'mensaje': plantilla})


# ─── Cálculos Laborales ─────────────────────────────────────────────

@login_required
def calculo_laboral(request, pk):
    """
    Vista de cálculos laborales para un expediente.
    Crea el CalculoLaboral si no existe, recalcula si es necesario.
    """
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    # Obtener o crear el cálculo
    calculo, created = CalculoLaboral.objects.get_or_create(expediente=expediente)

    if request.method == 'POST':
        form = CalculoLaboralForm(request.POST, instance=calculo)
        if form.is_valid():
            calculo = form.save(commit=False)
            # Recalcular con los parámetros actualizados
            recalcular_calculo(calculo)
            calculo.save()

            registrar_movimiento(
                expediente=expediente,
                usuario=request.user,
                accion='actualizacion',
                detalle='Cálculo laboral recalculado'
            )
            messages.success(request, '✅ Cálculo laboral recalculado correctamente.')
            return redirect('calculo_laboral', pk=expediente.pk)
        else:
            messages.error(request, 'Corrige los errores del formulario.')
    else:
        # Si es nuevo o el cálculo está vacío, recalcular automáticamente
        if created or calculo.total == 0:
            recalcular_calculo(calculo)
            calculo.save()
        form = CalculoLaboralForm(instance=calculo)

    # Obtener configuración legal activa
    config_legal = LegalConfig.get_active()

    # Verificar datos suficientes para calcular
    cliente = expediente.cliente
    datos_completos = all([
        cliente.fecha_ingreso,
        cliente.fecha_salida,
        cliente.salario and cliente.salario > 0,
    ])

    return render(request, 'expedientes/calculo_laboral.html', {
        'form': form,
        'expediente': expediente,
        'calculo': calculo,
        'config_legal': config_legal,
        'datos_completos': datos_completos,
        'simulacion_form': SimulacionForm(),
        'ESTADO_COLORS': ESTADO_COLORS,
        'es_nuevo': created,
    })


@login_required
def simulacion_rapida(request):
    """
    Simulación rápida de prestaciones vía AJAX/HTMX.
    Recibe datos por GET, devuelve HTML parcial con resultados.
    """
    if request.method == 'POST':
        form = SimulacionForm(request.POST)
        if form.is_valid():
            resultado = simular(
                salario=form.cleaned_data['salario'],
                fecha_ingreso=form.cleaned_data['fecha_ingreso'],
                fecha_salida=form.cleaned_data['fecha_salida'],
                periodo_pago=form.cleaned_data['periodo_pago'],
            )
            return render(request, 'expedientes/_simulacion_resultado.html', {
                'resultado': resultado,
            })
        else:
            return render(request, 'expedientes/_simulacion_resultado.html', {
                'resultado': {'success': False, 'error': 'Corrige los datos de la simulación'},
            })

    return render(request, 'expedientes/_simulacion_resultado.html', {
        'resultado': None,
    })


# ─── Solicitud de Conciliación ──────────────────────────────────────────

@login_required
def solicitud_conciliacion(request, pk):
    """Vista para llenar/editar el Formato para Iniciar la Solicitud de Conciliación."""
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)
    solicitud, created = SolicitudConciliacion.objects.get_or_create(expediente=expediente)

    if request.method == 'POST':
        form = SolicitudConciliacionForm(request.POST, instance=solicitud)
        if form.is_valid():
            form.save()
            registrar_movimiento(
                expediente=expediente,
                usuario=request.user,
                accion='actualizacion',
                detalle='Solicitud de conciliación actualizada'
            )
            messages.success(request, 'Solicitud de conciliación guardada correctamente.')
            return redirect('expediente_detail', pk=expediente.pk)
        else:
            messages.error(request, 'Corrige los errores del formulario.')
    else:
        form = SolicitudConciliacionForm(instance=solicitud)

    return render(request, 'expedientes/solicitud_conciliacion_form.html', {
        'form': form,
        'expediente': expediente,
        'solicitud': solicitud,
        'es_nueva': created,
        'ESTADO_COLORS': ESTADO_COLORS,
    })


# ─── Generador de Demanda (Editor WYSIWYG + Word) ───────────────────────

@login_required
def _puede_generar_documentos(request):
    """
    Helper: verifica si el usuario puede generar documentos legales.
    
    Tienen permiso:
    - Admin y Superadmin por defecto (su rol lo permite)
    - Cualquier usuario con el flag 'puede_generar_documentos' activado
      en su perfil (permiso granular por usuario)
    """
    if not hasattr(request.user, 'profile'):
        return False
    perfil = request.user.profile
    return perfil.rol in ['admin', 'superadmin'] or perfil.puede_generar_documentos


@login_required
def demanda_editor(request, pk):
    """
    Editor WYSIWYG de la demanda laboral.
    Muestra el contenido generado en un editor Quill.js para que
    el usuario pueda editarlo antes de descargar.
    Incluye selector de plantillas para diferentes tipos de despido
    y machotes (plantillas desde casos reales).
    """
    if not _puede_generar_documentos(request):
        messages.error(request, 'No tienes permiso para generar documentos legales.')
        return redirect('dashboard_asesor')
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    try:
        contenido_html = generar_demanda_html(expediente)
        # Generar todas las plantillas predefinidas para el selector
        plantillas = {}
        for tipo_key in PLANTILLAS_INFO:
            plantillas[tipo_key] = generar_demanda_html(expediente, tipo_despido_override=tipo_key)

        # Cargar machotes de demanda desde la BD (importados desde .docx)
        machotes_demanda = _get_machotes_queryset().filter(
            categoria='demanda',
        )

        # Reemplazar marcadores en cada machote con datos del expediente
        machotes_renderizados = []
        for m in machotes_demanda:
            try:
                contenido = reemplazar_marcadores(m.contenido_html, expediente)
                machotes_renderizados.append({
                    'id': m.pk,
                    'nombre': m.nombre,
                    'descripcion': m.descripcion,
                    'tipo_despido': m.tipo_despido,
                    'tipo_despido_display': m.get_tipo_despido_display() if m.tipo_despido else '',
                    'icono': m.icono,
                    'favorito': m.favorito,
                    'html': contenido,
                })
            except Exception:
                # Si falla el reemplazo de marcadores, incluir el HTML original
                machotes_renderizados.append({
                    'id': m.pk,
                    'nombre': m.nombre,
                    'descripcion': m.descripcion,
                    'tipo_despido': m.tipo_despido,
                    'tipo_despido_display': m.get_tipo_despido_display() if m.tipo_despido else '',
                    'icono': m.icono,
                    'favorito': m.favorito,
                    'html': m.contenido_html,
                })

    except Exception as e:
        logger.exception("Error generando HTML de demanda para expediente %s", expediente.numero)
        messages.error(request, 'Error al cargar el editor de demanda.')
        return redirect('expediente_detail', pk=expediente.pk)

    return render(request, 'expedientes/demanda_editor.html', {
        'expediente': expediente,
        'contenido_html': contenido_html,
        'plantillas': plantillas,
        'plantillas_info': PLANTILLAS_INFO,
        'machotes_demanda': machotes_renderizados,
        'tipo_actual': expediente.tipo_despido or 'injustificado',
        'ESTADO_COLORS': ESTADO_COLORS,
    })


@login_required
def demanda_descargar(request, pk):
    """
    Recibe el HTML editado desde el editor Quill.js y genera
    el documento Word para descarga.
    """
    if not _puede_generar_documentos(request):
        messages.error(request, 'No tienes permiso para descargar documentos legales.')
        return redirect('dashboard_asesor')
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    if request.method != 'POST':
        return redirect('demanda_editor', pk=expediente.pk)

    contenido_html = request.POST.get('contenido', '')
    if not contenido_html:
        messages.error(request, 'No hay contenido para generar la demanda.')
        return redirect('demanda_editor', pk=expediente.pk)

    try:
        from .demanda_generator import html_a_docx

        doc = generar_demanda_word(expediente, desde_cero=False)
        html_a_docx(contenido_html, doc)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        nombre_archivo = f"Demanda_{expediente.numero}_{re.sub(r'[^\w\-]', '_', expediente.cliente.nombre)}.docx"
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

        doc.save(response)

        registrar_movimiento(
            expediente=expediente,
            usuario=request.user,
            accion='actualizacion',
            detalle='Documento de demanda laboral generado desde el editor'
        )

        return response

    except Exception as e:
        logger.exception("Error generando demanda editada para expediente %s", expediente.numero)
        messages.error(request, 'Error al generar la demanda. Intenta de nuevo.')
        return redirect('demanda_editor', pk=expediente.pk)


@login_required
def generar_demanda(request, pk):
    """
    Genera y descarga un documento Word de Demanda Laboral
    con los datos del expediente y cálculos integrados.
    (Versión directa sin edición)
    """
    if not _puede_generar_documentos(request):
        messages.error(request, 'No tienes permiso para generar documentos legales.')
        return redirect('dashboard_asesor')
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    try:
        doc = generar_demanda_word(expediente)

        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        nombre_archivo = f"Demanda_{expediente.numero}_{re.sub(r'[^\w\-]', '_', expediente.cliente.nombre)}.docx"
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

        doc.save(response)

        registrar_movimiento(
            expediente=expediente,
            usuario=request.user,
            accion='actualizacion',
            detalle='Documento de demanda laboral generado y descargado'
        )

        return response

    except Exception as e:
        logger.exception("Error generando demanda para expediente %s", expediente.numero)
        messages.error(request, 'Error al generar la demanda. Intenta de nuevo o contacta al administrador.')
        return redirect('expediente_detail', pk=expediente.pk)


# ─── Transferencias de Casos ────────────────────────────────────────────

def _notificar_transferencia(solicitud, asesor_anterior, nuevo_asesor, admin_user, comentario=''):
    """
    Envía notificaciones por email al asesor anterior y al nuevo asesor
    cuando una transferencia es aprobada.
    """
    expediente = solicitud.expediente
    cliente = expediente.cliente

    asesor_anterior_nombre = asesor_anterior.get_full_name() or asesor_anterior.username
    nuevo_asesor_nombre = nuevo_asesor.get_full_name() or nuevo_asesor.username
    admin_nombre = admin_user.get_full_name() or admin_user.username

    # ─── Email al asesor anterior (se le fue aprobada la transferencia) ───
    asunto_anterior = f'✅ Transferencia aprobada — {expediente.numero} ({cliente.nombre})'
    mensaje_anterior = f"""
Hola {asesor_anterior_nombre},

Tu solicitud de transferencia para el caso {expediente.numero} ha sido APROBADA.

📋 Caso: {expediente.numero} — {cliente.nombre}
🔄 Asignado a: {nuevo_asesor_nombre}
📝 Motivo: {solicitud.motivo}
✅ Aprobado por: {admin_nombre}
"""
    if comentario:
        mensaje_anterior += f"\n💬 Comentario: {comentario}"
    mensaje_anterior += f"""

El caso ya no aparece en tu bandeja. El nuevo asesor se encargará del seguimiento.

Saludos,
Sistema de Gestión de Conciliacion Laboral Tijuana
"""

    # ─── Email al nuevo asesor (se le asignó un nuevo caso) ───
    asunto_nuevo = f'📋 Nuevo caso asignado — {expediente.numero} ({cliente.nombre})'
    mensaje_nuevo = f"""
Hola {nuevo_asesor_nombre},

Se te ha asignado un nuevo caso.

📋 Caso: {expediente.numero} — {cliente.nombre}
📌 Estado: {expediente.get_estado_display()}
👤 Asesor anterior: {asesor_anterior_nombre}
📝 Motivo de transferencia: {solicitud.motivo}
"""
    if comentario:
        mensaje_nuevo += f"💬 Comentario: {comentario}\n"
    mensaje_nuevo += f"""

Por favor revisa el expediente para dar continuidad al caso.

Saludos,
Sistema de Gestión de Conciliacion Laboral Tijuana
"""

    # Enviar emails
    fallos = []

    if asesor_anterior.email:
        try:
            send_mail(
                subject=asunto_anterior,
                message=mensaje_anterior,
                from_email=None,  # usa DEFAULT_FROM_EMAIL
                recipient_list=[asesor_anterior.email],
                fail_silently=False,
            )
        except Exception as e:
            logger.warning(f"Error enviando notificación a {asesor_anterior.email}: {e}")
            fallos.append(asesor_anterior.email)
    else:
        fallos.append(f"{asesor_anterior_nombre} (sin email)")

    if nuevo_asesor.email:
        try:
            send_mail(
                subject=asunto_nuevo,
                message=mensaje_nuevo,
                from_email=None,
                recipient_list=[nuevo_asesor.email],
                fail_silently=False,
            )
        except Exception as e:
            logger.warning(f"Error enviando notificación a {nuevo_asesor.email}: {e}")
            fallos.append(nuevo_asesor.email)
    else:
        fallos.append(f"{nuevo_asesor_nombre} (sin email)")

    return fallos


@login_required
def solicitar_transferencia(request, pk):
    """
    Muestra el formulario para solicitar la transferencia de un expediente.
    Solo el asesor asignado puede solicitar la transferencia.
    """
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    # Validar que solo el asesor asignado pueda solicitar transferencia
    if request.user != expediente.asesor:
        messages.error(request, 'Solo el asesor asignado al caso puede solicitar una transferencia.')
        return redirect('expediente_detail', pk=expediente.pk)

    # Validar que no tenga una solicitud pendiente
    if expediente.solicitudes_transferencia.filter(estado='pendiente').exists():
        messages.warning(request, 'Ya existe una solicitud de transferencia pendiente para este expediente.')
        return redirect('expediente_detail', pk=expediente.pk)

    asesores_disponibles = User.objects.filter(
        profile__rol='asesor',
        is_active=True
    ).exclude(pk=request.user.pk).order_by('first_name')

    return render(request, 'expedientes/transferencia_form.html', {
        'expediente': expediente,
        'asesores_disponibles': asesores_disponibles,
        'ESTADO_COLORS': ESTADO_COLORS,
    })


@login_required
def enviar_solicitud_transferencia(request, pk):
    """
    Procesa el formulario de solicitud de transferencia.
    """
    if request.method != 'POST':
        return redirect('solicitar_transferencia', pk=pk)

    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    if request.user != expediente.asesor:
        messages.error(request, 'No autorizado.')
        return redirect('expediente_detail', pk=expediente.pk)

    motivo = request.POST.get('motivo', '').strip()
    asesor_destino_id = request.POST.get('asesor_destino')

    if not motivo:
        messages.error(request, 'Debes explicar el motivo de la transferencia.')
        return redirect('solicitar_transferencia', pk=expediente.pk)

    asesor_destino = None
    if asesor_destino_id:
        try:
            asesor_destino = User.objects.get(
                pk=asesor_destino_id,
                profile__rol='asesor',
                is_active=True
            )
        except User.DoesNotExist:
            messages.warning(request, 'El asesor seleccionado no está disponible. La solicitud quedará abierta para que administración asigne.')

    SolicitudTransferencia.objects.create(
        expediente=expediente,
        solicitante=request.user,
        asesor_destino=asesor_destino,
        motivo=motivo,
    )

    registrar_movimiento(
        expediente=expediente,
        usuario=request.user,
        accion='actualizacion',
        detalle=f'Solicitud de transferencia enviada. Motivo: {motivo[:100]}'
    )

    messages.success(request, '✅ Solicitud de transferencia enviada. El área administrativa revisará y asignará un nuevo asesor.')
    return redirect('expediente_detail', pk=expediente.pk)


@login_required
def gestionar_transferencias(request):
    """
    Panel de gestión de transferencias para administradores.
    Muestra todas las solicitudes pendientes y el historial reciente.
    """
    if not hasattr(request.user, 'profile') or request.user.profile.rol not in ['admin', 'superadmin']:
        messages.error(request, 'No tienes permisos para gestionar transferencias.')
        return redirect('dashboard_asesor')

    pendientes = SolicitudTransferencia.objects.filter(
        estado='pendiente'
    ).select_related(
        'expediente', 'expediente__cliente',
        'solicitante', 'asesor_destino'
    ).order_by('-created_at')

    historial = SolicitudTransferencia.objects.exclude(
        estado='pendiente'
    ).select_related(
        'expediente', 'expediente__cliente',
        'solicitante', 'resuelto_por', 'asesor_asignado'
    ).order_by('-fecha_resolucion')[:20]

    asesores_disponibles = User.objects.filter(
        profile__rol='asesor',
        is_active=True
    ).order_by('first_name')

    return render(request, 'expedientes/transferencias_gestion.html', {
        'pendientes': pendientes,
        'historial': historial,
        'asesores_disponibles': asesores_disponibles,
        'ESTADO_COLORS': ESTADO_COLORS,
    })


@login_required
def aprobar_transferencia(request, pk):
    """
    Admin aprueba la transferencia y reasigna el expediente a un nuevo asesor.
    """
    if not hasattr(request.user, 'profile') or request.user.profile.rol not in ['admin', 'superadmin']:
        messages.error(request, 'No tienes permisos para aprobar transferencias.')
        return redirect('dashboard_asesor')

    solicitud = get_object_or_404(SolicitudTransferencia, pk=pk, estado='pendiente')
    expediente = solicitud.expediente

    if request.method == 'POST':
        nuevo_asesor_id = request.POST.get('nuevo_asesor')
        comentario = request.POST.get('comentario', '').strip()

        if not nuevo_asesor_id:
            messages.error(request, 'Debes seleccionar un asesor para reasignar el caso.')
            return redirect('gestionar_transferencias')

        try:
            nuevo_asesor = User.objects.get(
                pk=nuevo_asesor_id,
                profile__rol='asesor',
                is_active=True
            )
        except User.DoesNotExist:
            messages.error(request, 'El asesor seleccionado no es válido.')
            return redirect('gestionar_transferencias')

        # Reasignar el expediente
        asesor_anterior = expediente.asesor
        expediente.asesor = nuevo_asesor
        expediente.save()

        # Marcar solicitud como aprobada
        solicitud.estado = 'aprobada'
        solicitud.resuelto_por = request.user
        solicitud.asesor_asignado = nuevo_asesor
        solicitud.comentario_admin = comentario
        solicitud.fecha_resolucion = timezone.now()
        solicitud.save()

        registrar_movimiento(
            expediente=expediente,
            usuario=request.user,
            accion='actualizacion',
            detalle=(f'Expediente transferido de {asesor_anterior.get_full_name() or asesor_anterior.username} '
                     f'a {nuevo_asesor.get_full_name() or nuevo_asesor.username}. '
                     f'Motivo: {solicitud.motivo[:100]}')
        )

        # Notificar a los asesores por email
        fallos = _notificar_transferencia(
            solicitud=solicitud,
            asesor_anterior=asesor_anterior,
            nuevo_asesor=nuevo_asesor,
            admin_user=request.user,
            comentario=comentario,
        )

        # Crear notificaciones internas
        enlace = reverse('expediente_detail', kwargs={'pk': expediente.pk})

        _crear_notificacion(
            usuario=asesor_anterior,
            titulo=f'✅ Transferencia aprobada — {expediente.numero}',
            mensaje=f'Tu solicitud de transferencia para {expediente.cliente.nombre} fue aprobada. '
                    f'El caso fue reasignado a {nuevo_asesor.get_full_name() or nuevo_asesor.username}.',
            tipo='transferencia',
            link=enlace,
        )
        _crear_notificacion(
            usuario=nuevo_asesor,
            titulo=f'📋 Nuevo caso asignado — {expediente.numero}',
            mensaje=f'Se te ha asignado el caso de {expediente.cliente.nombre}. '
                    f'Asesor anterior: {asesor_anterior.get_full_name() or asesor_anterior.username}.',
            tipo='transferencia',
            link=enlace,
        )

        msg = f'✅ Transferencia aprobada. Caso {expediente.numero} reasignado a {nuevo_asesor.get_full_name() or nuevo_asesor.username}.'
        if fallos:
            msg += f' ⚠️ No se pudo notificar a: {", ".join(fallos)}'
        messages.success(request, msg)
        return redirect('gestionar_transferencias')

    return redirect('gestionar_transferencias')


@login_required
def rechazar_transferencia(request, pk):
    """
    Admin rechaza la solicitud de transferencia.
    """
    if not hasattr(request.user, 'profile') or request.user.profile.rol not in ['admin', 'superadmin']:
        messages.error(request, 'No tienes permisos para rechazar transferencias.')
        return redirect('dashboard_asesor')

    solicitud = get_object_or_404(SolicitudTransferencia, pk=pk, estado='pendiente')
    expediente = solicitud.expediente

    if request.method == 'POST':
        comentario = request.POST.get('comentario', '').strip()

        solicitud.estado = 'rechazada'
        solicitud.resuelto_por = request.user
        solicitud.comentario_admin = comentario
        solicitud.fecha_resolucion = timezone.now()
        solicitud.save()

        registrar_movimiento(
            expediente=expediente,
            usuario=request.user,
            accion='actualizacion',
            detalle=f'Solicitud de transferencia rechazada. Motivo: {comentario or "Sin comentario"}'
        )

        # Notificar al solicitante que su transferencia fue rechazada
        _crear_notificacion(
            usuario=solicitud.solicitante,
            titulo=f'❌ Transferencia rechazada — {expediente.numero}',
            mensaje=f'Tu solicitud de transferencia para {expediente.cliente.nombre} fue rechazada. '
                    f'Comentario: {comentario or "Sin comentario"}.',
            tipo='transferencia',
            link=reverse('expediente_detail', kwargs={'pk': expediente.pk}),
        )

        messages.warning(request, f'Solicitud de transferencia rechazada para el caso {expediente.numero}.')
        return redirect('gestionar_transferencias')

    return redirect('gestionar_transferencias')


@login_required
def cancelar_transferencia(request, pk):
    """
    El asesor cancela su propia solicitud de transferencia (si aún está pendiente).
    """
    solicitud = get_object_or_404(
        SolicitudTransferencia,
        pk=pk,
        solicitante=request.user,
        estado='pendiente'
    )
    expediente = solicitud.expediente

    if request.method == 'POST':
        solicitud.estado = 'cancelada'
        solicitud.fecha_resolucion = timezone.now()
        solicitud.save()

        registrar_movimiento(
            expediente=expediente,
            usuario=request.user,
            accion='actualizacion',
            detalle='Solicitud de transferencia cancelada por el solicitante.'
        )

        messages.info(request, 'Solicitud de transferencia cancelada.')
        return redirect('expediente_detail', pk=expediente.pk)

    return redirect('expediente_detail', pk=expediente.pk)


# ─── Notificaciones ──────────────────────────────────────────────────

@require_POST
def marcar_notificacion_leida(request, pk):
    """Marca una notificación como leída."""
    noti = get_object_or_404(Notificacion, pk=pk, usuario=request.user)
    noti.leida = True
    noti.save()
    return JsonResponse({'success': True})


@require_POST
def marcar_todas_leidas(request):
    """Marca todas las notificaciones del usuario como leídas."""
    Notificacion.objects.filter(usuario=request.user, leida=False).update(leida=True)
    return JsonResponse({'success': True, 'count': 0})


def _crear_notificacion(usuario, titulo, mensaje='', tipo='sistema', link=''):
    """Helper para crear una notificación."""
    return Notificacion.objects.create(
        usuario=usuario,
        titulo=titulo,
        mensaje=mensaje,
        tipo=tipo,
        link=link,
    )


# ─── Toggle favorito machote ───────────────────────────────────────────

@login_required
@require_POST
def toggle_machote_favorito(request, machote_id):
    """
    Marca/desmarca un machote como favorito.
    Los favoritos aparecen primero en el editor de demanda.
    """
    machote = get_object_or_404(Machote, pk=machote_id, activo=True)
    machote.favorito = not machote.favorito
    machote.save(update_fields=['favorito'])

    return JsonResponse({
        'success': True,
        'favorito': machote.favorito,
        'machote_id': machote.pk,
    })


# ─── Machotes / Generador de Documentos ────────────────────────────────


@login_required
def machotes_listar(request, pk):
    """
    Lista los machotes disponibles para un expediente.
    Filtra por tipo de despido si el expediente tiene uno asignado.
    """
    if not _puede_generar_documentos(request):
        messages.error(request, 'No tienes permiso para acceder al generador de documentos.')
        return redirect('dashboard_asesor')
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    # Obtener machotes activos
    machotes_qs = Machote.objects.filter(activo=True)

    # Si el expediente tiene tipo de despido, preferir machotes de ese tipo
    if expediente.tipo_despido:
        machotes_qs = machotes_qs.filter(
            Q(tipo_despido=expediente.tipo_despido) | Q(tipo_despido__isnull=True)
        )

    machotes = machotes_qs.order_by('categoria', 'orden', 'nombre')

    # Agrupar por categoría
    from collections import defaultdict
    machotes_agrupados = defaultdict(list)
    for m in machotes:
        machotes_agrupados[m.get_categoria_display()].append(m)

    # Estadísticas de completitud de datos
    campos = get_datos_faltantes(expediente)
    completitud_stats = get_completitud_stats(campos)

    return render(request, 'expedientes/machotes_listar.html', {
        'expediente': expediente,
        'machotes_agrupados': dict(machotes_agrupados),
        'completitud_stats': completitud_stats,
        'ESTADO_COLORS': ESTADO_COLORS,
    })


@login_required
def documento_preparar(request, pk, machote_id):
    """
    Página de preparación previa a generar un documento desde un machote.
    Muestra checklist de datos completos/faltantes, resumen de cálculos,
    y permite generar el documento cuando el usuario esté listo.
    """
    if not _puede_generar_documentos(request):
        messages.error(request, 'No tienes permiso para acceder al generador de documentos.')
        return redirect('dashboard_asesor')
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)
    machote = get_object_or_404(Machote, pk=machote_id, activo=True)

    # Obtener datos de completitud
    calculo = calcular_desde_expediente(expediente)
    campos = get_datos_faltantes(expediente, calculo)
    stats = get_completitud_stats(campos)

    # Obtener marcadores disponibles en el machote
    marcadores_disponibles = machote.get_marcadores_disponibles()

    # Vista previa de los valores que se inyectarán
    preview = get_marcadores(expediente, calculo)
    # Solo mostrar los marcadores que están en el machote
    preview_filtrado = {}
    for m in marcadores_disponibles:
        if m in preview:
            preview_filtrado[m] = preview[m]

    return render(request, 'expedientes/documento_preparar.html', {
        'expediente': expediente,
        'machote': machote,
        'campos': campos,
        'stats': stats,
        'calculo': calculo,
        'marcadores_disponibles': marcadores_disponibles,
        'preview': preview_filtrado or preview,
        'ESTADO_COLORS': ESTADO_COLORS,
    })


@login_required
def machotes_generar(request, pk, machote_id):
    """
    Genera un documento HTML a partir de un machote,
    reemplazando los marcadores con datos del expediente.
    """
    if not _puede_generar_documentos(request):
        messages.error(request, 'No tienes permiso para generar documentos legales.')
        return redirect('dashboard_asesor')
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)
    machote = get_object_or_404(Machote, pk=machote_id, activo=True)

    # Reemplazar marcadores con datos reales (usando módulo especializado)
    contenido = reemplazar_marcadores(machote.contenido_html, expediente)

    # Registrar movimiento
    registrar_movimiento(
        expediente=expediente,
        usuario=request.user,
        accion='actualizacion',
        detalle=f'Documento generado desde machote: {machote.nombre}'
    )

    messages.success(request, f'✅ Documento generado desde "{machote.nombre}". Puedes editarlo antes de descargar.')

    return redirect('machotes_editor', pk=expediente.pk, machote_id=machote.pk)


@login_required
def machotes_editor(request, pk, machote_id):
    """
    Editor WYSIWYG para documentos generados desde machotes.
    Similar al demanda_editor pero para cualquier tipo de machote.
    """
    if not _puede_generar_documentos(request):
        messages.error(request, 'No tienes permiso para editar documentos legales.')
        return redirect('dashboard_asesor')
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)
    machote = get_object_or_404(Machote, pk=machote_id, activo=True)

    # Generar contenido con datos (usando módulo especializado)
    contenido_html = reemplazar_marcadores(machote.contenido_html, expediente)

    return render(request, 'expedientes/machotes_editor.html', {
        'expediente': expediente,
        'machote': machote,
        'contenido_html': contenido_html,
        'machotes_relacionados': Machote.objects.filter(
            activo=True, categoria=machote.categoria
        ).exclude(pk=machote.pk)[:4],
        'ESTADO_COLORS': ESTADO_COLORS,
    })


@login_required
def machotes_descargar(request, pk, machote_id):
    """
    Descarga el documento editado (desde el editor WYSIWYG) como .docx
    """
    if not _puede_generar_documentos(request):
        messages.error(request, 'No tienes permiso para descargar documentos legales.')
        return redirect('dashboard_asesor')
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)
    machote = get_object_or_404(Machote, pk=machote_id, activo=True)

    if request.method != 'POST':
        return redirect('machotes_editor', pk=expediente.pk, machote_id=machote.pk)

    contenido_html = request.POST.get('contenido', '')
    if not contenido_html:
        messages.error(request, 'No hay contenido para generar el documento.')
        return redirect('machotes_editor', pk=expediente.pk, machote_id=machote.pk)

    try:
        doc = generar_demanda_word(expediente, desde_cero=False)
        html_a_docx(contenido_html, doc)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        nombre_safe = re.sub(r'[^\w\-]', '_', machote.nombre)[:40]
        nombre_archivo = f"{nombre_safe}_{expediente.numero}.docx"
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

        doc.save(response)

        registrar_movimiento(
            expediente=expediente,
            usuario=request.user,
            accion='actualizacion',
            detalle=f'Documento descargado desde machote: {machote.nombre}'
        )

        return response

    except Exception as e:
        logger.exception("Error generando documento desde machote %s", machote.nombre)
        messages.error(request, 'Error al generar el documento. Intenta de nuevo.')
        return redirect('machotes_editor', pk=expediente.pk, machote_id=machote.pk)


# ─── Reportes Administrativos ─────────────────────────────────────────────

@login_required
def reportes_admin(request):
    """Vista de reportes administrativos con estadísticas detalladas."""
    if not hasattr(request.user, 'profile') or request.user.profile.rol not in ['admin', 'superadmin']:
        return redirect('dashboard_asesor')

    qs = Expediente.objects.all()

    # Reporte: casos por asesor
    asesores = User.objects.filter(profile__rol='asesor').annotate(
        total_casos=Count('expediente'),
        casos_activos=Count('expediente', filter=~Q(expediente__estado='cerrado')),
        casos_cerrados=Count('expediente', filter=Q(expediente__estado='cerrado')),
        convenios=Count('expediente', filter=Q(expediente__estado='convenio')),
        demandas=Count('expediente', filter=Q(expediente__estado='demanda')),
    ).order_by('-total_casos')

    # Audiencias semanales
    hoy = timezone.now()
    inicio_semana = hoy - timezone.timedelta(days=hoy.weekday())
    fin_semana = inicio_semana + timezone.timedelta(days=6)
    audiencias_semanales = qs.filter(
        fecha_audiencia__date__gte=inicio_semana.date(),
        fecha_audiencia__date__lte=fin_semana.date(),
    ).order_by('fecha_audiencia')

    context = {
        'asesores': asesores,
        'audiencias_semanales': audiencias_semanales,
        'total_casos': qs.count(),
        'total_cerrados': qs.filter(estado='cerrado').count(),
        'total_convenios': qs.filter(estado='convenio').count(),
        'total_demandas': qs.filter(estado='demanda').count(),
        'inicio_semana': inicio_semana,
        'fin_semana': fin_semana,
        'ESTADO_COLORS': ESTADO_COLORS,
    }

    return render(request, 'expedientes/reportes_admin.html', context)


# ═══════════════════════════════════════════════════════════════════════════
#  Automatización de Conciliación — ASÍNCRONA (threading)
# ═══════════════════════════════════════════════════════════════════════════

def _ejecutar_conciliacion_en_hilo(task_id):
    """
    Ejecuta la automatización de conciliación en un hilo separado.

    Esto evita que el HTTP request se bloquee durante los 30-90 segundos
    que tarda Playwright en navegar y llenar el formulario del portal.
    """
    from django.db import close_old_connections

    close_old_connections()
    logger.info(f'[Hilo] Iniciando tarea de conciliación #{task_id}')

    try:
        # Marcar como ejecutando
        task = TareaConciliacion.objects.get(pk=task_id)
        task.estado = 'ejecutando'
        task.save(update_fields=['estado'])

        close_old_connections()

        # Ejecutar la automatización (esto puede tomar 30-90 segundos)
        resultado = enviar_conciliacion(
            task.expediente,
            usuario=task.usuario,
            headless=True,  # Siempre headless en background
        )

        close_old_connections()

        # Recargar la tarea desde la BD (por si cambió entre tanto)
        task.refresh_from_db()

        if resultado.success:
            task.estado = 'completado'
            task.folio = resultado.folio or ''
            task.pdf_path = resultado.pdf_path or ''
        else:
            task.estado = 'fallido'
            task.error = resultado.error or 'Error desconocido al enviar al portal'

        task.detalle = resultado.detalle or ''
        if resultado.screenshots:
            task.screenshots_json = json.dumps(resultado.screenshots)
        else:
            task.screenshots_json = ''
        task.completed_at = timezone.now()
        task.save()

        logger.info(f'[Hilo] Tarea #{task_id} completada: {task.estado}')

    except Exception as e:
        close_old_connections()
        logger.exception(f'[Hilo] Error en tarea de conciliación #{task_id}')
        try:
            task = TareaConciliacion.objects.get(pk=task_id)
            task.estado = 'fallido'
            task.error = f'{type(e).__name__}: {str(e)}'
            task.completed_at = timezone.now()
            task.save(update_fields=['estado', 'error', 'completed_at'])
        except:
            pass


@login_required
def enviar_conciliacion_automation(request, pk):
    """
    Vista que inicia el envío automático al portal de conciliación.

    AHORA ES ASÍNCRONA: crea una tarea en BD, la ejecuta en un hilo
    separado y redirige a una página de progreso para evitar timeouts.
    """
    expediente = get_object_or_404(get_expedientes_queryset(request.user), pk=pk)

    # Verificar que el expediente tenga datos mínimos
    if not expediente.cliente.curp:
        messages.error(request, 'El cliente debe tener CURP para enviar la solicitud de conciliación.')
        return redirect('expediente_detail', pk=expediente.pk)

    if not expediente.cliente.telefono:
        messages.error(request, 'El cliente debe tener teléfono registrado.')
        return redirect('expediente_detail', pk=expediente.pk)

    if request.method == 'POST':
        modo = request.POST.get('modo', 'automatico')

        # Crear la tarea en la BD
        task = TareaConciliacion.objects.create(
            expediente=expediente,
            usuario=request.user,
            estado='pendiente',
            modo=modo,
        )

        # Iniciar la ejecución en un hilo separado
        hilo = threading.Thread(
            target=_ejecutar_conciliacion_en_hilo,
            args=(task.pk,),
            daemon=True,
        )
        hilo.start()

        messages.info(request, '🚀 Iniciando envío automático al portal de conciliación...')
        return redirect('conciliacion_procesando', task_pk=task.pk)

    # GET: mostrar página de confirmación
    return render(request, 'expedientes/conciliacion_confirmar.html', {
        'expediente': expediente,
        'ESTADO_COLORS': ESTADO_COLORS,
    })


@login_required
def conciliacion_estado(request, task_pk):
    """
    API JSON: retorna el estado actual de una tarea de conciliación.
    Usado por la página de progreso para polling vía JavaScript.
    """
    task = get_object_or_404(TareaConciliacion.objects.select_related('expediente'), pk=task_pk)

    # Asegurar que el usuario tenga acceso al expediente
    expedientes_qs = get_expedientes_queryset(request.user)
    if not expedientes_qs.filter(pk=task.expediente.pk).exists():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    data = {
        'task_id': task.pk,
        'estado': task.estado,
        'expediente_id': task.expediente_id,
        'expediente_numero': task.expediente.numero,
        'folio': task.folio,
        'error': task.error,
        'detalle': task.detalle,
        'created_at': task.created_at.isoformat() if task.created_at else None,
        'completed_at': task.completed_at.isoformat() if task.completed_at else None,
    }

    return JsonResponse(data)


@login_required
def conciliacion_procesando(request, task_pk):
    """
    Página de progreso: muestra una animación mientras se ejecuta
    la automatización en segundo plano y redirige cuando termina.
    """
    task = get_object_or_404(TareaConciliacion, pk=task_pk)

    # Asegurar que el usuario tenga acceso al expediente
    expedientes_qs = get_expedientes_queryset(request.user)
    if not expedientes_qs.filter(pk=task.expediente.pk).exists():
        messages.error(request, 'No tienes acceso a este expediente.')
        return redirect('dashboard_redirect')

    return render(request, 'expedientes/conciliacion_procesando.html', {
        'task': task,
        'expediente': task.expediente,
        'ESTADO_COLORS': ESTADO_COLORS,
    })


@login_required
def reintentar_conciliacion(request, task_pk):
    """
    Reintenta una tarea de conciliación fallida.
    Crea una nueva tarea para el mismo expediente y la ejecuta en un hilo.
    """
    task_original = get_object_or_404(TareaConciliacion, pk=task_pk)

    # Verificar que el usuario tenga acceso al expediente
    expedientes_qs = get_expedientes_queryset(request.user)
    expediente = get_object_or_404(expedientes_qs, pk=task_original.expediente.pk)

    # Verificar datos mínimos
    if not expediente.cliente.curp:
        messages.error(request, 'El cliente debe tener CURP para enviar la solicitud de conciliación.')
        return redirect('expediente_detail', pk=expediente.pk)

    if not expediente.cliente.telefono:
        messages.error(request, 'El cliente debe tener teléfono registrado.')
        return redirect('expediente_detail', pk=expediente.pk)

    # Evitar reintentos duplicados si ya hay una tarea en progreso
    tarea_existente = expediente.tareas_conciliacion.filter(
        estado__in=['pendiente', 'ejecutando']
    ).first()
    if tarea_existente:
        messages.warning(request, 'Ya hay un envío en progreso para este expediente. Espera a que termine.')
        return redirect('conciliacion_procesando', task_pk=tarea_existente.pk)

    # Crear nueva tarea
    task = TareaConciliacion.objects.create(
        expediente=expediente,
        usuario=request.user,
        estado='pendiente',
        modo=task_original.modo or 'automatico',
    )

    # Iniciar la ejecución en un hilo separado
    hilo = threading.Thread(
        target=_ejecutar_conciliacion_en_hilo,
        args=(task.pk,),
        daemon=True,
    )
    hilo.start()

    messages.info(request, '🚀 Reintentando envío automático al portal de conciliación...')
    return redirect('conciliacion_procesando', task_pk=task.pk)
