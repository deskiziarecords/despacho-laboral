import logging
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User
from django.db.models import Sum, Q, Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.utils import timezone
from django.urls import reverse_lazy
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib import messages

from .models import SettlementPayment, Expense, Commission, CashMovement, Office, Payroll, Partner, WorkWeek, PartnerLoan, Agreement, Honorario
from .forms import CashMovementForm, PartnerForm, WorkWeekForm, PartnerLoanForm, AgreementForm, HonorarioForm
from expedientes.models import Expediente


logger = logging.getLogger(__name__)


def _calcular_flujo_mensual(oficina_id=None):
    """
    Helper: calcula el flujo de caja mensual de los últimos 12 meses.
    Retorna dict con labels, ingresos, gastos, utilidad.
    """
    hoy = timezone.now()
    meses_labels = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
                    'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    meses = []
    ingresos_meses = []
    gastos_meses = []
    utilidad_meses = []

    for i in range(11, -1, -1):
        if i == 0:
            inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            fin_mes = hoy
        else:
            mid = (hoy.replace(day=1) - timedelta(days=1)).replace(day=1)
            for _ in range(i - 1):
                mid = (mid - timedelta(days=1)).replace(day=1)
            inicio_mes = mid
            ultimo_dia = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
            fin_mes = ultimo_dia if ultimo_dia < hoy else hoy

        fd_m = inicio_mes.date() if inicio_mes else None
        fh_m = fin_mes.date() if fin_mes else None

        def filtrar_mes(qs, campo='fecha'):
            q = qs
            if oficina_id:
                q = q.filter(oficina_id=oficina_id)
            if fd_m:
                q = q.filter(**{f'{campo}__gte': fd_m})
            if fh_m:
                q = q.filter(**{f'{campo}__lte': fh_m})
            return q

        ing_m = (
            (filtrar_mes(SettlementPayment.objects.all()).aggregate(t=Sum('monto'))['t'] or 0)
            + (filtrar_mes(CashMovement.objects.filter(tipo='ingreso')).aggregate(t=Sum('monto'))['t'] or 0)
        )
        gas_m = (
            (filtrar_mes(Expense.objects.all()).aggregate(t=Sum('monto'))['t'] or 0)
            + (filtrar_mes(CashMovement.objects.filter(tipo='egreso')).aggregate(t=Sum('monto'))['t'] or 0)
            + (filtrar_mes(Payroll.objects.all(), campo='fecha_pago').aggregate(t=Sum('total_pagado'))['t'] or 0)
        )
        uti_m = ing_m - gas_m

        meses.append(meses_labels[inicio_mes.month - 1][:3])
        ingresos_meses.append(float(ing_m))
        gastos_meses.append(float(gas_m))
        utilidad_meses.append(float(uti_m))

    return {
        'labels': meses,
        'ingresos': ingresos_meses,
        'gastos': gastos_meses,
        'utilidad': utilidad_meses,
    }


class AdminOrSuperOnlyMixin(UserPassesTestMixin):
    """Restringe acceso a administrativos y superadmin."""
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.rol in ['admin', 'superadmin', 'finanzas']


class DashboardFinancieroView(LoginRequiredMixin, AdminOrSuperOnlyMixin, TemplateView):
    """
    Dashboard financiero con resumen de ingresos vs gastos vs utilidad.
    Visible solo para administradores y superadmin.
    """
    template_name = 'finanzas/dashboard_financiero.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        hoy = timezone.now()
        mes_inicio = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        mes_fin = (mes_inicio + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
        anio_inicio = hoy.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

        # ─── Fechas para filtros ───────────────────────────────────────
        periodo = self.request.GET.get('periodo', 'mes')
        if periodo == 'mes':
            fecha_desde = mes_inicio
            fecha_hasta = mes_fin
            etiqueta_periodo = f"Mes actual ({hoy.strftime('%B %Y').capitalize()})"
        elif periodo == 'anio':
            fecha_desde = anio_inicio
            fecha_hasta = hoy
            etiqueta_periodo = f"Año {hoy.year}"
        else:  # todo
            fecha_desde = None
            fecha_hasta = None
            etiqueta_periodo = "Histórico total"

        # ─── Filtro por oficina ─────────────────────────────────────────
        oficina_id = self.request.GET.get('oficina')

        def filtrar_por_oficina(qs):
            if oficina_id:
                return qs.filter(oficina_id=oficina_id)
            return qs

        def filtrar_por_fecha(qs, campo_fecha='fecha'):
            if fecha_desde:
                qs = qs.filter(**{f'{campo_fecha}__gte': fecha_desde.date()})
            if fecha_hasta:
                qs = qs.filter(**{f'{campo_fecha}__lte': fecha_hasta.date()})
            return qs

        # ─── 1. Ingresos (SettlementPayment + CashMovement ingresos) ──
        ingresos_pagos = filtrar_por_oficina(
            filtrar_por_fecha(SettlementPayment.objects.all())
        ).aggregate(total=Sum('monto'))['total'] or 0

        ingresos_caja = filtrar_por_oficina(
            filtrar_por_fecha(CashMovement.objects.filter(tipo='ingreso'))
        ).aggregate(total=Sum('monto'))['total'] or 0

        total_ingresos = ingresos_pagos + ingresos_caja

        # ─── 2. Gastos (Expense + CashMovement egresos) ────────────────
        gastos_expense = filtrar_por_oficina(
            filtrar_por_fecha(Expense.objects.all())
        ).aggregate(total=Sum('monto'))['total'] or 0

        gastos_caja = filtrar_por_oficina(
            filtrar_por_fecha(CashMovement.objects.filter(tipo='egreso'))
        ).aggregate(total=Sum('monto'))['total'] or 0

        # Gastos de nómina (Payroll)
        gastos_nomina = filtrar_por_oficina(
            filtrar_por_fecha(Payroll.objects.all(), campo_fecha='fecha_pago')
        ).aggregate(total=Sum('total_pagado'))['total'] or 0

        total_gastos = gastos_expense + gastos_caja + gastos_nomina

        # ─── 3. Utilidad ────────────────────────────────────────────────
        utilidad = total_ingresos - total_gastos

        # ─── 4. Resumen por oficina ────────────────────────────────────
        oficinas = Office.objects.filter(activa=True)
        resumen_oficinas = []
        for of in oficinas:
            ing_of = (
                filtrar_por_fecha(of.settlementpayment_set.all()).aggregate(total=Sum('monto'))['total'] or 0
                + filtrar_por_fecha(of.cashmovement_set.filter(tipo='ingreso')).aggregate(total=Sum('monto'))['total'] or 0
            )
            gas_of = (
                filtrar_por_fecha(of.expense_set.all()).aggregate(total=Sum('monto'))['total'] or 0
                + filtrar_por_fecha(of.cashmovement_set.filter(tipo='egreso')).aggregate(total=Sum('monto'))['total'] or 0
                + filtrar_por_fecha(of.payroll_set.all(), campo_fecha='fecha_pago').aggregate(total=Sum('total_pagado'))['total'] or 0
            )
            uti_of = ing_of - gas_of

            # Comisiones pagadas en la oficina
            com_pagadas = filtrar_por_fecha(
                Commission.objects.filter(oficina=of, estado='pagada')
            ).aggregate(total=Sum('monto_comision'))['total'] or 0

            resumen_oficinas.append({
                'oficina': of,
                'ingresos': ing_of,
                'gastos': gas_of,
                'utilidad': uti_of,
                'comisiones_pagadas': com_pagadas,
            })

        context['resumen_oficinas'] = resumen_oficinas

        # ─── 5. Ingresos y Gastos por categoría ────────────────────────
        # Categorías de gastos
        gastos_por_categoria = filtrar_por_oficina(
            filtrar_por_fecha(Expense.objects.values('categoria'))
        ).annotate(total=Sum('monto')).order_by('-total')

        context['gastos_por_categoria'] = gastos_por_categoria

        # Formas de pago más usadas
        formas_pago = filtrar_por_oficina(
            filtrar_por_fecha(SettlementPayment.objects.values('forma_pago'))
        ).annotate(total=Sum('monto')).order_by('-total')

        context['formas_pago'] = formas_pago

        # ─── 6. Productividad por asesor (desde expedientes) ───────────
        asesores = User.objects.filter(profile__rol='asesor', is_active=True).annotate(
            total_casos=Count('expediente'),
            casos_activos=Count('expediente', filter=~Q(expediente__estado='cerrado')),
            convenios=Count('expediente', filter=Q(expediente__estado='convenio')),
        ).order_by('-total_casos')

        asesores_data = []
        for asesor in asesores:
            # Comisiones del asesor en el período
            com_asesor = filtrar_por_fecha(
                Commission.objects.filter(asesor=asesor, estado='pagada')
            ).aggregate(total=Sum('monto_comision'))['total'] or 0

            # Monto recuperado (convenios)
            monto_recuperado = filtrar_por_fecha(
                Expediente.objects.filter(asesor=asesor), campo_fecha='created_at'
            ).aggregate(total=Sum('monto_convenio'))['total'] or 0

            asesores_data.append({
                'asesor': asesor,
                'total_casos': asesor.total_casos,
                'casos_activos': asesor.casos_activos,
                'convenios': asesor.convenios,
                'monto_recuperado': monto_recuperado,
                'comisiones': com_asesor,
            })

        context['asesores'] = asesores_data

        # ─── 7. Flujo de caja mensual (12 meses) ───────────────────
        flujo = _calcular_flujo_mensual(oficina_id=oficina_id)
        context['meses_labels'] = flujo['labels']
        context['ingresos_meses'] = flujo['ingresos']
        context['gastos_meses'] = flujo['gastos']
        context['utilidad_meses'] = flujo['utilidad']

        # ─── 8. Movimientos de caja recientes ──────────────────────────
        context['caja_reciente'] = CashMovement.objects.select_related(
            'oficina', 'registrado_por'
        ).order_by('-fecha', '-created_at')[:15]

        # ─── Contexto global ────────────────────────────────────────────
        context['total_ingresos'] = total_ingresos
        context['total_gastos'] = total_gastos
        context['utilidad'] = utilidad
        context['utilidad_positivo'] = utilidad >= 0
        context['ingresos_pagos'] = ingresos_pagos
        context['ingresos_caja'] = ingresos_caja
        context['gastos_expense'] = gastos_expense
        context['gastos_caja'] = gastos_caja
        context['gastos_nomina'] = gastos_nomina
        context['periodo'] = periodo
        context['etiqueta_periodo'] = etiqueta_periodo
        context['oficina_seleccionada'] = int(oficina_id) if oficina_id else None
        context['oficinas'] = oficinas
        context['hoy'] = hoy

        return context


# ─── CRUD Movimientos de Caja ────────────────────────────────────────────────


class CashMovementListView(LoginRequiredMixin, AdminOrSuperOnlyMixin, ListView):
    """
    Listado de movimientos de caja con filtros por oficina, tipo, fechas y búsqueda.
    """
    model = CashMovement
    template_name = 'finanzas/cashmovement_list.html'
    paginate_by = 25
    context_object_name = 'movimientos'

    def get_queryset(self):
        qs = CashMovement.objects.select_related('oficina', 'registrado_por').order_by('-fecha', '-created_at')

        tipo = self.request.GET.get('tipo', '')
        oficina_id = self.request.GET.get('oficina', '')
        categoria = self.request.GET.get('categoria', '')
        fecha_desde = self.request.GET.get('fecha_desde', '')
        fecha_hasta = self.request.GET.get('fecha_hasta', '')
        q = self.request.GET.get('q', '')

        if tipo:
            qs = qs.filter(tipo=tipo)
        if oficina_id:
            qs = qs.filter(oficina_id=oficina_id)
        if categoria:
            qs = qs.filter(categoria=categoria)
        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)
        if q:
            qs = qs.filter(descripcion__icontains=q) | qs.filter(referencia__icontains=q)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['oficinas'] = Office.objects.filter(activa=True)
        context['categorias_ingreso'] = CashMovement.CATEGORIA_INGRESO_CHOICES
        context['categorias_egreso'] = CashMovement.CATEGORIA_EGRESO_CHOICES
        context['filtros'] = {k: v for k, v in self.request.GET.items() if v and k != 'page'}

        # Totales del listado filtrado
        qs = self.get_queryset()
        totales = qs.aggregate(
            total_ingresos=Sum('monto', filter=Q(tipo='ingreso')),
            total_egresos=Sum('monto', filter=Q(tipo='egreso')),
        )
        context['total_ingresos_filtro'] = totales['total_ingresos'] or 0
        context['total_egresos_filtro'] = totales['total_egresos'] or 0
        context['total_movimientos'] = qs.count()

        return context


class CashMovementCreateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, CreateView):
    """
    Crear un nuevo movimiento de caja (ingreso o egreso).
    """
    model = CashMovement
    form_class = CashMovementForm
    template_name = 'finanzas/cashmovement_form.html'
    success_url = reverse_lazy('cashmovement_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        # Pre-fill with default values from query params (for quick-add from dashboard)
        initial = {}
        tipo = self.request.GET.get('tipo')
        if tipo in ['ingreso', 'egreso']:
            initial['tipo'] = tipo
        oficina_id = self.request.GET.get('oficina')
        if oficina_id:
            try:
                initial['oficina'] = int(oficina_id)
            except (ValueError, TypeError):
                pass
        if initial:
            kwargs['initial'] = initial
        return kwargs

    def form_valid(self, form):
        form.instance.registrado_por = self.request.user
        messages.success(self.request, f'✅ Movimiento de {"ingreso" if form.instance.tipo == "ingreso" else "egreso"} registrado correctamente.')
        return super().form_valid(form)


class CashMovementUpdateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, UpdateView):
    """
    Editar un movimiento de caja existente.
    """
    model = CashMovement
    form_class = CashMovementForm
    template_name = 'finanzas/cashmovement_form.html'
    success_url = reverse_lazy('cashmovement_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, '✅ Movimiento de caja actualizado correctamente.')
        return super().form_valid(form)


class CashMovementDeleteView(LoginRequiredMixin, AdminOrSuperOnlyMixin, DeleteView):
    """
    Eliminar un movimiento de caja.
    """
    model = CashMovement
    template_name = 'finanzas/cashmovement_confirm_delete.html'
    success_url = reverse_lazy('cashmovement_list')
    context_object_name = 'movimiento'

    def delete(self, request, *args, **kwargs):
        movimiento = self.get_object()
        messages.success(request, f'✅ Movimiento de {"ingreso" if movimiento.tipo == "ingreso" else "egreso"} eliminado correctamente.')
        return super().delete(request, *args, **kwargs)


# ─── CRUD Socios ───────────────────────────────────────────────────────────


class PartnerListView(LoginRequiredMixin, AdminOrSuperOnlyMixin, ListView):
    model = Partner
    template_name = 'finanzas/partner_list.html'
    context_object_name = 'partners'
    paginate_by = 20

    def get_queryset(self):
        qs = Partner.objects.all().order_by('-activo', 'nombre')
        q = self.request.GET.get('q', '')
        if q:
            qs = qs.filter(nombre__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtros'] = {k: v for k, v in self.request.GET.items() if v}
        return context


class PartnerCreateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, CreateView):
    model = Partner
    form_class = PartnerForm
    template_name = 'finanzas/partner_form.html'
    success_url = reverse_lazy('partner_list')

    def form_valid(self, form):
        messages.success(self.request, f'✅ Socio {form.instance.nombre} creado correctamente.')
        return super().form_valid(form)


class PartnerUpdateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, UpdateView):
    model = Partner
    form_class = PartnerForm
    template_name = 'finanzas/partner_form.html'
    success_url = reverse_lazy('partner_list')

    def form_valid(self, form):
        messages.success(self.request, f'✅ Socio {form.instance.nombre} actualizado correctamente.')
        return super().form_valid(form)


class PartnerDetailView(LoginRequiredMixin, AdminOrSuperOnlyMixin, DetailView):
    model = Partner
    template_name = 'finanzas/partner_detail.html'
    context_object_name = 'partner'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['prestamos_origen'] = self.object.prestamos_origen.select_related(
            'socio_destino', 'registrado_por'
        ).order_by('-fecha')[:20]
        context['prestamos_destino'] = self.object.prestamos_destino.select_related(
            'socio_origen', 'registrado_por'
        ).order_by('-fecha')[:20]
        return context


# ─── CRUD Semanas de Trabajo ───────────────────────────────────────────────


class WorkWeekListView(LoginRequiredMixin, AdminOrSuperOnlyMixin, ListView):
    model = WorkWeek
    template_name = 'finanzas/workweek_list.html'
    context_object_name = 'weeks'
    paginate_by = 20

    def get_queryset(self):
        qs = WorkWeek.objects.all().order_by('-fecha_inicio')
        estado = self.request.GET.get('estado', '')
        if estado:
            qs = qs.filter(estado=estado)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtros'] = {k: v for k, v in self.request.GET.items() if v}
        return context


class WorkWeekCreateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, CreateView):
    model = WorkWeek
    form_class = WorkWeekForm
    template_name = 'finanzas/workweek_form.html'
    success_url = reverse_lazy('workweek_list')

    def get_initial(self):
        initial = super().get_initial()
        from datetime import date, timedelta
        hoy = date.today()
        inicio = hoy - timedelta(days=hoy.weekday())
        fin = inicio + timedelta(days=6)
        initial['numero'] = hoy.isocalendar()[1]
        initial['fecha_inicio'] = inicio
        initial['fecha_fin'] = fin
        return initial

    def form_valid(self, form):
        messages.success(self.request, f'✅ Semana {form.instance.numero} creada correctamente.')
        return super().form_valid(form)


class WorkWeekUpdateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, UpdateView):
    model = WorkWeek
    form_class = WorkWeekForm
    template_name = 'finanzas/workweek_form.html'
    success_url = reverse_lazy('workweek_list')

    def form_valid(self, form):
        messages.success(self.request, f'✅ Semana {form.instance.numero} actualizada correctamente.')
        return super().form_valid(form)


# ─── CRUD Convenios (Agreements) ─────────────────────────────────────────


class AgreementListView(LoginRequiredMixin, AdminOrSuperOnlyMixin, ListView):
    model = Agreement
    template_name = 'finanzas/agreement_list.html'
    context_object_name = 'agreements'
    paginate_by = 20

    def get_queryset(self):
        qs = Agreement.objects.select_related(
            'cliente', 'oficina', 'responsable', 'creado_por'
        ).order_by('-fecha', '-created_at')

        estado = self.request.GET.get('estado', '')
        oficina_id = self.request.GET.get('oficina', '')
        q = self.request.GET.get('q', '')

        if estado:
            qs = qs.filter(estado=estado)
        if oficina_id:
            qs = qs.filter(oficina_id=oficina_id)
        if q:
            qs = qs.filter(
                Q(cliente__nombre__icontains=q) |
                Q(empresa__icontains=q) |
                Q(notas__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['oficinas'] = Office.objects.filter(activa=True)
        context['filtros'] = {k: v for k, v in self.request.GET.items() if v and k != 'page'}
        context['estado_choices'] = Agreement.ESTADO_CHOICES

        # Totales
        qs = self.get_queryset()
        totales = qs.aggregate(
            total_monto=Sum('monto_convenio'),
            total_honorarios=Sum('honorarios'),
            pendientes=Count('pk', filter=Q(estado='pendiente')),
            pagados=Count('pk', filter=Q(estado='pagado')),
        )
        context['total_monto'] = totales['total_monto'] or 0
        context['total_honorarios'] = totales['total_honorarios'] or 0
        context['pendientes_count'] = totales['pendientes'] or 0
        context['pagados_count'] = totales['pagados'] or 0
        context['total_agreements'] = qs.count()

        return context


class AgreementCreateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, CreateView):
    model = Agreement
    form_class = AgreementForm
    template_name = 'finanzas/agreement_form.html'
    success_url = reverse_lazy('agreement_list')

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, f'✅ Convenio de {form.instance.cliente.nombre} creado correctamente.')
        return super().form_valid(form)

    def get_initial(self):
        initial = super().get_initial()
        initial['fecha'] = timezone.now().date()
        initial['responsable'] = self.request.user
        return initial


class AgreementUpdateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, UpdateView):
    model = Agreement
    form_class = AgreementForm
    template_name = 'finanzas/agreement_form.html'
    success_url = reverse_lazy('agreement_list')

    def form_valid(self, form):
        messages.success(self.request, f'✅ Convenio de {form.instance.cliente.nombre} actualizado correctamente.')
        return super().form_valid(form)


class AgreementDetailView(LoginRequiredMixin, AdminOrSuperOnlyMixin, DetailView):
    model = Agreement
    template_name = 'finanzas/agreement_detail.html'
    context_object_name = 'agreement'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['honorarios'] = self.object.honorario_set.select_related(
            'registrado_por'
        ).order_by('-created_at')
        context['total_honorarios_pendientes'] = self.object.honorarios_pendientes
        context['total_honorarios_pagados'] = self.object.honorarios_pagados
        context['pagos'] = SettlementPayment.objects.filter(
            cliente=self.object.cliente
        ).order_by('-fecha')[:10]
        return context


# ─── CRUD Honorarios ───────────────────────────────────────────────────────


class HonorarioCreateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, CreateView):
    model = Honorario
    form_class = HonorarioForm
    template_name = 'finanzas/honorario_form.html'
    success_url = reverse_lazy('agreement_list')

    def get_initial(self):
        initial = super().get_initial()
        convenio_id = self.request.GET.get('convenio')
        if convenio_id:
            try:
                initial['convenio'] = int(convenio_id)
            except (ValueError, TypeError):
                pass
        return initial

    def form_valid(self, form):
        form.instance.registrado_por = self.request.user
        messages.success(
            self.request,
            f'✅ Honorario de {form.instance.porcentaje}% creado correctamente. '
            f'Monto calculado: ${form.instance.monto_calculado:,.2f}'
        )
        return super().form_valid(form)

    def get_success_url(self):
        if self.object:
            return reverse_lazy('agreement_detail', kwargs={'pk': self.object.convenio.pk})
        return reverse_lazy('agreement_list')


class HonorarioUpdateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, UpdateView):
    model = Honorario
    form_class = HonorarioForm
    template_name = 'finanzas/honorario_form.html'

    def form_valid(self, form):
        messages.success(self.request, f'✅ Honorario actualizado correctamente.')
        return super().form_valid(form)

    def get_success_url(self):
        if self.object:
            return reverse_lazy('agreement_detail', kwargs={'pk': self.object.convenio.pk})
        return reverse_lazy('agreement_list')


# ─── CRUD Préstamos entre Socios ───────────────────────────────────────────


class PartnerLoanListView(LoginRequiredMixin, AdminOrSuperOnlyMixin, ListView):
    model = PartnerLoan
    template_name = 'finanzas/partnerloan_list.html'
    context_object_name = 'loans'
    paginate_by = 20

    def get_queryset(self):
        qs = PartnerLoan.objects.select_related(
            'socio_origen', 'socio_destino', 'registrado_por'
        ).order_by('-fecha', '-created_at')

        estado = self.request.GET.get('estado', '')
        socio = self.request.GET.get('socio', '')

        if estado:
            qs = qs.filter(estado=estado)
        if socio:
            qs = qs.filter(
                Q(socio_origen_id=socio) | Q(socio_destino_id=socio)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtros'] = {k: v for k, v in self.request.GET.items() if v}
        context['socios'] = Partner.objects.filter(activo=True)

        # Totales
        qs = self.get_queryset()
        context['total_pendiente'] = qs.filter(estado='pendiente').aggregate(
            t=models.Sum('monto')
        )['t'] or 0
        context['total_pagado'] = qs.filter(estado='pagado').aggregate(
            t=models.Sum('monto')
        )['t'] or 0
        return context


class PartnerLoanCreateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, CreateView):
    model = PartnerLoan
    form_class = PartnerLoanForm
    template_name = 'finanzas/partnerloan_form.html'
    success_url = reverse_lazy('partnerloan_list')

    def form_valid(self, form):
        form.instance.registrado_por = self.request.user
        messages.success(self.request, '✅ Préstamo registrado correctamente.')
        return super().form_valid(form)


class PartnerLoanUpdateView(LoginRequiredMixin, AdminOrSuperOnlyMixin, UpdateView):
    model = PartnerLoan
    form_class = PartnerLoanForm
    template_name = 'finanzas/partnerloan_form.html'
    success_url = reverse_lazy('partnerloan_list')

    def form_valid(self, form):
        messages.success(self.request, '✅ Préstamo actualizado correctamente.')
        return super().form_valid(form)


@login_required
def api_flujo_mensual(request):
    """
    API JSON: retorna datos del flujo de caja mensual filtrado por oficina.
    """
    if not hasattr(request.user, 'profile') or request.user.profile.rol not in ['admin', 'superadmin', 'finanzas']:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    oficina_id = request.GET.get('oficina')
    data = _calcular_flujo_mensual(oficina_id=oficina_id)
    return JsonResponse(data)


@login_required
def exportar_dashboard_financiero_excel(request):
    """
    Exporta el Dashboard Financiero a un archivo Excel con múltiples hojas.
    """
    if not hasattr(request.user, 'profile') or request.user.profile.rol not in ['admin', 'superadmin', 'finanzas']:
        return redirect('dashboard_asesor')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ─── Estilos compartidos ──────────────────────────────────────
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        title_font = Font(bold=True, size=12, color='1F2937')
        green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

        def estilo_header(ws, num_cols):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

        def auto_ancho(ws, min_width=12, max_width=40):
            for col_cells in ws.columns:
                length = max((len(str(c.value or '')) for c in col_cells), default=0)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 3, min_width), max_width)

        # ═══════════════════════════════════════════════════════════════
        # HOJA 1: Resumen Global
        # ═══════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Resumen Global'

        # Título
        hoy = timezone.now()
        ws1.cell(row=1, column=1, value=f'Dashboard Financiero — {hoy.strftime("%B %Y").capitalize()}').font = title_font
        ws1.merge_cells('A1:F1')
        ws1.cell(row=2, column=1, value=f'Generado: {hoy.strftime("%d/%m/%Y %H:%M")}').font = Font(size=9, color='666666')
        ws1.merge_cells('A2:F2')

        # Totales
        periodo = request.GET.get('periodo', 'mes')
        oficina_id = request.GET.get('oficina')

        # Reusar lógica de fechas del dashboard (código inline simplificado)
        hoy_dt = timezone.now()
        mes_inicio = hoy_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        mes_fin = (mes_inicio + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
        anio_inicio = hoy_dt.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

        if periodo == 'mes':
            fd, fh = mes_inicio, mes_fin
        elif periodo == 'anio':
            fd, fh = anio_inicio, hoy_dt
        else:
            fd, fh = None, None

        def filtrar_oficina(qs):
            return qs.filter(oficina_id=oficina_id) if oficina_id else qs

        def filtrar_fecha(qs, campo='fecha'):
            if fd:
                qs = qs.filter(**{f'{campo}__gte': fd.date()})
            if fh:
                qs = qs.filter(**{f'{campo}__lte': fh.date()})
            return qs

        i_pagos = filtrar_oficina(filtrar_fecha(SettlementPayment.objects.all())).aggregate(t=Sum('monto'))['t'] or 0
        i_caja = filtrar_oficina(filtrar_fecha(CashMovement.objects.filter(tipo='ingreso'))).aggregate(t=Sum('monto'))['t'] or 0
        total_ing = i_pagos + i_caja

        g_exp = filtrar_oficina(filtrar_fecha(Expense.objects.all())).aggregate(t=Sum('monto'))['t'] or 0
        g_caja = filtrar_oficina(filtrar_fecha(CashMovement.objects.filter(tipo='egreso'))).aggregate(t=Sum('monto'))['t'] or 0
        g_nom = filtrar_oficina(filtrar_fecha(Payroll.objects.all(), campo='fecha_pago')).aggregate(t=Sum('total_pagado'))['t'] or 0
        total_gas = g_exp + g_caja + g_nom
        utilidad = total_ing - total_gas

        row = 4
        ws1.cell(row=row, column=1, value='Métrica').font = Font(bold=True)
        ws1.cell(row=row, column=2, value='Monto').font = Font(bold=True)
        estilo_header(ws1, 2)
        row += 1

        for label, val, fill in [
            ('💰 Ingresos por Pagos', i_pagos, green_fill),
            ('💰 Ingresos por Caja', i_caja, green_fill),
            ('💰 Total Ingresos', total_ing, green_fill),
            ('💸 Gastos Operativos', g_exp, red_fill),
            ('💸 Gastos de Caja', g_caja, red_fill),
            ('💸 Gastos de Nómina', g_nom, red_fill),
            ('💸 Total Gastos', total_gas, red_fill),
            ('📊 Utilidad', utilidad, green_fill if utilidad >= 0 else red_fill),
        ]:
            ws1.cell(row=row, column=1, value=label).border = thin_border
            c = ws1.cell(row=row, column=2, value=float(val))
            c.number_format = '$#,##0.00'
            c.border = thin_border
            if fill:
                c.fill = fill
                ws1.cell(row=row, column=1).fill = fill
            row += 1

        # Margen
        row += 1
        ws1.cell(row=row, column=1, value='Margen de Utilidad').font = Font(bold=True)
        ws1.cell(row=row, column=1).border = thin_border
        c = ws1.cell(row=row, column=2, value=float(utilidad / total_ing * 100) if total_ing > 0 else 0)
        c.number_format = '0.0"%"'
        c.border = thin_border

        # ─── Resumen por Oficina ────────────────────────────────────
        row += 2
        ws1.cell(row=row, column=1, value='Resumen por Oficina').font = title_font
        ws1.merge_cells(f'A{row}:F{row}')
        row += 1
        headers_of = ['Oficina', 'Ingresos', 'Gastos', 'Utilidad', 'Margen %', 'Comisiones Pagadas']
        for col, h in enumerate(headers_of, 1):
            ws1.cell(row=row, column=col, value=h)
        estilo_header(ws1, len(headers_of))
        row += 1

        oficinas = Office.objects.filter(activa=True)
        for of in oficinas:
            ing_of = (
                filtrar_fecha(of.settlementpayment_set.all()).aggregate(t=Sum('monto'))['t'] or 0
                + filtrar_fecha(of.cashmovement_set.filter(tipo='ingreso')).aggregate(t=Sum('monto'))['t'] or 0
            )
            gas_of = (
                filtrar_fecha(of.expense_set.all()).aggregate(t=Sum('monto'))['t'] or 0
                + filtrar_fecha(of.cashmovement_set.filter(tipo='egreso')).aggregate(t=Sum('monto'))['t'] or 0
                + filtrar_fecha(of.payroll_set.all(), campo='fecha_pago').aggregate(t=Sum('total_pagado'))['t'] or 0
            )
            com_of = filtrar_fecha(Commission.objects.filter(oficina=of, estado='pagada')).aggregate(t=Sum('monto_comision'))['t'] or 0
            uti_of = ing_of - gas_of
            margen = float(uti_of / ing_of * 100) if ing_of > 0 else 0

            ws1.cell(row=row, column=1, value=of.nombre).border = thin_border
            ws1.cell(row=row, column=2, value=float(ing_of)).number_format = '$#,##0.00'
            ws1.cell(row=row, column=2).border = thin_border
            ws1.cell(row=row, column=3, value=float(gas_of)).number_format = '$#,##0.00'
            ws1.cell(row=row, column=3).border = thin_border
            ws1.cell(row=row, column=4, value=float(uti_of)).number_format = '$#,##0.00'
            ws1.cell(row=row, column=4).border = thin_border
            ws1.cell(row=row, column=5, value=margen).number_format = '0.0"%"'
            ws1.cell(row=row, column=5).border = thin_border
            ws1.cell(row=row, column=6, value=float(com_of)).number_format = '$#,##0.00'
            ws1.cell(row=row, column=6).border = thin_border
            row += 1

        auto_ancho(ws1)

        # ═══════════════════════════════════════════════════════════════
        # HOJA 2: Gastos por Categoría
        # ═══════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Gastos por Categoría')
        headers = ['Categoría', 'Total Gastado']
        for col, h in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=h)
        estilo_header(ws2, len(headers))

        gastos_cat = filtrar_oficina(filtrar_fecha(Expense.objects.values('categoria'))).annotate(t=Sum('monto')).order_by('-t')
        for i, g in enumerate(gastos_cat, 2):
            ws2.cell(row=i, column=1, value=g.get_categoria_display()).border = thin_border
            c = ws2.cell(row=i, column=2, value=float(g['t']))
            c.number_format = '$#,##0.00'
            c.border = thin_border
        auto_ancho(ws2)

        # ═══════════════════════════════════════════════════════════════
        # HOJA 3: Formas de Pago
        # ═══════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('Formas de Pago')
        headers = ['Forma de Pago', 'Total']
        for col, h in enumerate(headers, 1):
            ws3.cell(row=1, column=col, value=h)
        estilo_header(ws3, len(headers))

        formas = filtrar_oficina(filtrar_fecha(SettlementPayment.objects.values('forma_pago'))).annotate(t=Sum('monto')).order_by('-t')
        for i, f in enumerate(formas, 2):
            ws3.cell(row=i, column=1, value=f.get_forma_pago_display()).border = thin_border
            c = ws3.cell(row=i, column=2, value=float(f['t']))
            c.number_format = '$#,##0.00'
            c.border = thin_border
        auto_ancho(ws3)

        # ═══════════════════════════════════════════════════════════════
        # HOJA 4: Productividad por Asesor
        # ═══════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet('Asesores')
        headers = ['Asesor', 'Casos Totales', 'Casos Activos', 'Convenios', 'Monto Recuperado', 'Comisiones Pagadas']
        for col, h in enumerate(headers, 1):
            ws4.cell(row=1, column=col, value=h)
        estilo_header(ws4, len(headers))

        asesores = User.objects.filter(profile__rol='asesor', is_active=True).annotate(
            total_casos=Count('expediente'),
            casos_activos=Count('expediente', filter=~Q(expediente__estado='cerrado')),
            convenios=Count('expediente', filter=Q(expediente__estado='convenio')),
        ).order_by('-total_casos')

        for i, asesor in enumerate(asesores, 2):
            com = filtrar_fecha(Commission.objects.filter(asesor=asesor, estado='pagada')).aggregate(t=Sum('monto_comision'))['t'] or 0
            monto_rec = filtrar_fecha(Expediente.objects.filter(asesor=asesor), campo='created_at').aggregate(t=Sum('monto_convenio'))['t'] or 0

            ws4.cell(row=i, column=1, value=asesor.get_full_name() or asesor.username).border = thin_border
            ws4.cell(row=i, column=2, value=asesor.total_casos).border = thin_border
            ws4.cell(row=i, column=3, value=asesor.casos_activos).border = thin_border
            ws4.cell(row=i, column=4, value=asesor.convenios).border = thin_border
            c = ws4.cell(row=i, column=5, value=float(monto_rec))
            c.number_format = '$#,##0.00'
            c.border = thin_border
            c = ws4.cell(row=i, column=6, value=float(com))
            c.number_format = '$#,##0.00'
            c.border = thin_border
        auto_ancho(ws4)

        # ═══════════════════════════════════════════════════════════════
        # HOJA 5: Movimientos de Caja Recientes
        # ═══════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet('Movimientos de Caja')
        headers = ['Fecha', 'Oficina', 'Tipo', 'Categoría', 'Monto', 'Descripción', 'Registrado por']
        for col, h in enumerate(headers, 1):
            ws5.cell(row=1, column=col, value=h)
        estilo_header(ws5, len(headers))

        caja = CashMovement.objects.select_related('oficina', 'registrado_por').order_by('-fecha', '-created_at')[:50]
        for i, m in enumerate(caja, 2):
            ws5.cell(row=i, column=1, value=m.fecha.strftime('%d/%m/%Y')).border = thin_border
            ws5.cell(row=i, column=2, value=m.oficina.nombre).border = thin_border
            ws5.cell(row=i, column=3, value=m.get_tipo_display()).border = thin_border
            ws5.cell(row=i, column=4, value=m.get_categoria_display()).border = thin_border
            c = ws5.cell(row=i, column=5, value=float(m.monto))
            c.number_format = '$#,##0.00'
            c.border = thin_border
            ws5.cell(row=i, column=6, value=m.descripcion[:80]).border = thin_border
            ws5.cell(row=i, column=7, value=m.registrado_por.get_full_name() or m.registrado_por.username).border = thin_border
        auto_ancho(ws5)

        # ─── Response ──────────────────────────────────────────────────
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = hoy.strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename=dashboard_financiero_{timestamp}.xlsx'
        wb.save(response)
        return response

    except ImportError:
        return HttpResponse('openpyxl no está instalado. Ejecute: pip install openpyxl', status=500)
    except Exception as e:
        logger.exception("Error exportando dashboard financiero")
        return HttpResponse(f'Error al generar el Excel: {e}', status=500)
