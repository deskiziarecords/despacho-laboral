from django.shortcuts import render, redirect
from django.contrib.auth import login, logout
from django.contrib.auth.views import LoginView, LogoutView, PasswordResetView, PasswordResetDoneView, PasswordResetConfirmView, PasswordResetCompleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import TemplateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from io import StringIO
from django.contrib.auth.decorators import login_required
from .forms import LoginForm


class SuperadminOnlyMixin(UserPassesTestMixin):
    """Restringe acceso solo a superadmin."""
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.rol == 'superadmin'


class CustomLoginView(LoginView):
    form_class = LoginForm
    template_name = 'accounts/login.html'

    def get_success_url(self):
        user = self.request.user
        if hasattr(user, 'profile'):
            rol = user.profile.rol
            if rol == 'asesor':
                return reverse_lazy('dashboard_asesor')
            elif rol == 'finanzas':
                return reverse_lazy('dashboard_financiero')
            elif rol in ['admin', 'superadmin']:
                return reverse_lazy('dashboard_admin')
        return reverse_lazy('dashboard_asesor')

    def form_invalid(self, form):
        messages.error(self.request, 'Usuario o contraseña incorrectos.')
        return super().form_invalid(form)


class CustomLogoutView(LogoutView):
    next_page = reverse_lazy('login')


# ─── Password Reset ─────────────────────────────────────────────────────

class CustomPasswordResetView(PasswordResetView):
    template_name = 'accounts/password_reset.html'
    email_template_name = 'accounts/password_reset_email.html'
    subject_template_name = 'accounts/password_reset_subject.txt'
    success_url = reverse_lazy('password_reset_done')


class CustomPasswordResetDoneView(PasswordResetDoneView):
    template_name = 'accounts/password_reset_done.html'


class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = 'accounts/password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')


class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = 'accounts/password_reset_complete.html'


# ─── Dashboards ──────────────────────────────────────────────────────────

class DashboardAsesorView(LoginRequiredMixin, TemplateView):
    template_name = 'expedientes/dashboard_asesor.html'

    def get(self, request, *args, **kwargs):
        if not hasattr(request.user, 'profile') or request.user.profile.rol != 'asesor':
            return redirect('dashboard_admin')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from expedientes.models import Expediente
        user = self.request.user
        context['total_casos'] = Expediente.objects.filter(asesor=user).count()
        context['casos_activos'] = Expediente.objects.filter(asesor=user).exclude(estado='cerrado').count()
        context['casos_cerrados'] = Expediente.objects.filter(asesor=user, estado='cerrado').count()
        context['proximas_audiencias'] = Expediente.objects.filter(
            asesor=user,
            fecha_audiencia__isnull=False,
        ).exclude(estado='cerrado').order_by('fecha_audiencia')[:5]
        context['mis_casos'] = Expediente.objects.filter(asesor=user).order_by('-created_at')[:10]
        context['alertas'] = Expediente.objects.filter(asesor=user, prioridad='alta', estado__in=['nuevo', 'solicitud', 'audiencia']).count()
        context['casos_pendientes'] = Expediente.objects.filter(asesor=user, estado__in=['nuevo', 'solicitud']).count()
        return context


class DashboardAdminView(LoginRequiredMixin, TemplateView):
    template_name = 'expedientes/dashboard_admin.html'

    def get(self, request, *args, **kwargs):
        if hasattr(request.user, 'profile') and request.user.profile.rol == 'asesor':
            return redirect('dashboard_asesor')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from expedientes.models import Expediente, Movimiento
        from django.contrib.auth.models import User
        from django.db.models import Count, Sum, Q
        from django.utils import timezone

        qs = Expediente.objects.all()

        context['total_casos'] = qs.count()
        context['casos_activos'] = qs.exclude(estado='cerrado').count()
        context['casos_cerrados'] = qs.filter(estado='cerrado').count()
        context['convenios_count'] = qs.filter(estado='convenio').count()
        context['demandas_count'] = qs.filter(estado='demanda').count()

        montos = qs.aggregate(
            total_reclamado=Sum('monto_reclamado'),
            total_convenio=Sum('monto_convenio')
        )
        context['total_reclamado'] = montos['total_reclamado'] or 0
        context['total_convenio'] = montos['total_convenio'] or 0

        asesores = User.objects.filter(profile__rol='asesor').annotate(
            total_casos=Count('expediente'),
            casos_activos=Count('expediente', filter=Q(expediente__estado__in=['nuevo', 'solicitud', 'citatorio', 'audiencia', 'no_notificado', 'reprogramacion', 'convenio', 'sin_conciliacion', 'demanda'])),
            casos_cerrados=Count('expediente', filter=Q(expediente__estado='cerrado')),
            convenios=Count('expediente', filter=Q(expediente__estado='convenio')),
        )
        context['asesores'] = asesores

        context['proximas_audiencias'] = qs.filter(
            fecha_audiencia__isnull=False,
        ).exclude(estado='cerrado').order_by('fecha_audiencia')[:10]

        context['ultimos_movimientos'] = Movimiento.objects.select_related(
            'expediente', 'usuario'
        ).order_by('-created_at')[:15]

        context['casos_por_estado'] = {
            label: qs.filter(estado=key).count()
            for key, label in Expediente.ESTADO_CHOICES
        }

        return context


# ─── Superadmin Panel ─────────────────────────────────────────────────────

@login_required
def exportar_matriz_permisos_excel(request):
    """
    Exporta la matriz de permisos a un archivo Excel.
    """
    if not hasattr(request.user, 'profile') or request.user.profile.rol != 'superadmin':
        messages.error(request, 'No tienes permisos para exportar esta matriz.')
        return redirect('dashboard_asesor')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Matriz de Permisos'

        headers = [
            'Usuario', 'Nombre Completo', 'Email', 'Rol',
            '¿Puede generar documentos?', 'Staff Django', 'Superusuario',
            '¿Está activo?', 'Casos Asignados', 'Movimientos Realizados',
            'Último Acceso', 'Fecha de Creación'
        ]

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        usuarios = User.objects.all().select_related('profile').annotate(
            total_expedientes=Count('expediente'),
            total_movimientos=Count('movimiento'),
        ).order_by('profile__rol', 'username')

        for i, u in enumerate(usuarios, 2):
            perfil = u.profile if hasattr(u, 'profile') else None
            rol = perfil.get_rol_display() if perfil else 'Sin perfil'
            puede_docs = 'Sí' if (perfil and perfil.puede_generar_documentos) else 'No'
            es_staff = 'Sí' if u.is_staff else 'No'
            es_super = 'Sí' if u.is_superuser else 'No'
            activo = 'Sí' if u.is_active else 'No'
            ultimo_acceso = u.last_login.strftime('%d/%m/%Y %H:%M') if u.last_login else 'Nunca'
            fecha_creacion = u.date_joined.strftime('%d/%m/%Y') if u.date_joined else ''

            row_data = [
                u.username,
                u.get_full_name() or '—',
                u.email or '—',
                rol,
                puede_docs,
                es_staff,
                es_super,
                activo,
                u.total_expedientes,
                u.total_movimientos,
                ultimo_acceso,
                fecha_creacion,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

                # Color de fondo según rol
                if perfil:
                    if perfil.rol == 'superadmin':
                        cell.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
                    elif perfil.rol == 'admin':
                        cell.fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

                # Resaltar inactivos en rojo
                if not u.is_active:
                    cell.fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')

        # Ajustar ancho de columnas
        column_widths = [22, 30, 30, 16, 22, 14, 14, 14, 16, 20, 18, 16]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=matriz_permisos_{timezone.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    except ImportError:
        return HttpResponse('openpyxl no está instalado. Ejecute: pip install openpyxl', status=500)


class MatrizPermisosView(LoginRequiredMixin, SuperadminOnlyMixin, TemplateView):
    """
    Matriz completa de permisos de todos los usuarios del sistema.
    Muestra cada usuario con todos sus roles y permisos en formato tabular.
    """
    template_name = 'accounts/matriz_permisos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from django.db.models import Count

        usuarios = User.objects.all().select_related('profile').annotate(
            total_expedientes=Count('expediente'),
            total_movimientos=Count('movimiento'),
        ).order_by('profile__rol', 'username')

        permisos_data = []
        for u in usuarios:
            perfil = u.profile if hasattr(u, 'profile') else None

            permisos_data.append({
                'user': u,
                'username': u.username,
                'nombre_completo': u.get_full_name() or '—',
                'email': u.email or '—',
                'rol': perfil.rol if perfil else 'sin_perfil',
                'rol_display': perfil.get_rol_display() if perfil else '—',
                'puede_generar_documentos': perfil.puede_generar_documentos if perfil else False,
                'is_staff': u.is_staff,
                'is_active': u.is_active,
                'is_superuser': u.is_superuser,
                'ultimo_acceso': u.last_login,
                'fecha_union': u.date_joined,
                'total_expedientes': u.total_expedientes,
                'total_movimientos': u.total_movimientos,
            })

        # Resumen por rol
        resumen = {
            'total': len(usuarios),
            'superadmins': sum(1 for p in permisos_data if p['rol'] == 'superadmin'),
            'admins': sum(1 for p in permisos_data if p['rol'] == 'admin'),
            'asesores': sum(1 for p in permisos_data if p['rol'] == 'asesor'),
            'sin_perfil': sum(1 for p in permisos_data if p['rol'] == 'sin_perfil'),
            'activos': sum(1 for p in permisos_data if p['is_active']),
            'inactivos': sum(1 for p in permisos_data if not p['is_active']),
            'con_docs': sum(1 for p in permisos_data if p['puede_generar_documentos']),
            'staff': sum(1 for p in permisos_data if p['is_staff']),
        }

        # Auditoría reciente de cambios de permisos
        from .models import PermisoAuditLog

        audit_log = PermisoAuditLog.objects.select_related(
            'usuario_modificado', 'usuario_modificado__profile',
            'modificado_por'
        )[:50]

        # Resumen de actividad (últimos 30 días)
        hace_30_dias = timezone.now() - timedelta(days=30)
        activity = PermisoAuditLog.objects.filter(created_at__gte=hace_30_dias)
        total_cambios_30d = activity.count()
        cambios_por_tipo = {}
        for accion_key, accion_label in PermisoAuditLog.ACCION_CHOICES:
            cambios_por_tipo[accion_label] = activity.filter(accion=accion_key).count()

        context['permisos'] = permisos_data
        context['resumen'] = resumen
        context['audit_log'] = audit_log
        context['total_cambios_30d'] = total_cambios_30d
        context['cambios_por_tipo'] = cambios_por_tipo
        context['fecha_generacion'] = timezone.now()

        return context


@login_required
def cargar_datos_demo(request):
    """
    Carga los datos de demostración (usuarios de prueba + casos de ejemplo).
    Solo accesible para superadmin desde el panel.
    """
    # Permitir superadmin y admin (el admin de Railway tiene rol='asesor' hasta que
    # se redeploye con el fix de entrypoint.sh, así que también permitimos 'admin')
    if not hasattr(request.user, 'profile') or request.user.profile.rol not in ['superadmin', 'admin']:
        messages.error(request, 'No tienes permisos para cargar datos de demostración.')
        return redirect('dashboard_asesor')

    if request.method == 'POST':
        from django.core.management import call_command

        output = StringIO()

        # 1. Crear usuarios de prueba
        try:
            call_command('crear_usuarios_prueba', stdout=output, stderr=output)
        except Exception as e:
            output.write(f'\n⚠️ Error en crear_usuarios_prueba: {e}\n')

        # 2. Sembrar datos de prueba
        try:
            call_command('seed_datos', stdout=output, stderr=output)
        except Exception as e:
            output.write(f'\n⚠️ Error en seed_datos: {e}\n')

        # 3. Asegurar que el superadmin tenga rol correcto
        try:
            if hasattr(request.user, 'profile') and request.user.profile.rol != 'superadmin':
                request.user.profile.rol = 'superadmin'
                request.user.profile.save()
                output.write('>>> Perfil de superadmin actualizado correctamente.\n')
        except Exception as e:
            output.write(f'⚠️ Error actualizando perfil: {e}\n')

        resultado = output.getvalue()

        # Verificar counts finales
        from expedientes.models import Cliente, Expediente
        clientes = Cliente.objects.count()
        expedientes = Expediente.objects.count()

        messages.success(request, f'✅ Datos de demostración cargados: {clientes} clientes, {expedientes} expedientes.')

        return render(request, 'accounts/cargar_datos_demo.html', {
            'resultado': resultado,
            'clientes': clientes,
            'expedientes': expedientes,
            'titulo': '✅ Datos cargados exitosamente',
        })

    # GET: mostrar confirmación
    return render(request, 'accounts/cargar_datos_demo.html', {
        'resultado': '',
        'clientes': 0,
        'expedientes': 0,
        'titulo': 'Cargar datos de demostración',
    })


class SuperadminDashboardView(LoginRequiredMixin, SuperadminOnlyMixin, TemplateView):
    """Panel exclusivo para superadmin con auditoría completa del sistema."""
    template_name = 'accounts/superadmin_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from expedientes.models import Expediente, Movimiento, Cliente, Documento

        hoy = timezone.now()
        hace_7_dias = hoy - timedelta(days=7)
        hace_30_dias = hoy - timedelta(days=30)

        context['total_usuarios'] = User.objects.count()
        context['total_expedientes'] = Expediente.objects.count()
        context['total_clientes'] = Cliente.objects.count()
        context['total_documentos'] = Documento.objects.count()
        context['total_movimientos'] = Movimiento.objects.count()

        context['total_superadmins'] = User.objects.filter(profile__rol='superadmin').count()
        context['total_admins'] = User.objects.filter(profile__rol='admin').count()
        context['total_asesores'] = User.objects.filter(profile__rol='asesor').count()

        context['usuarios_activos_7d'] = User.objects.filter(
            last_login__gte=hace_7_dias
        ).count()
        context['usuarios_activos_30d'] = User.objects.filter(
            last_login__gte=hace_30_dias
        ).count()
        context['usuarios_sin_login'] = User.objects.filter(last_login__isnull=True).count()

        context['movimientos_7d'] = Movimiento.objects.filter(created_at__gte=hace_7_dias).count()
        context['movimientos_30d'] = Movimiento.objects.filter(created_at__gte=hace_30_dias).count()
        context['expedientes_creados_7d'] = Expediente.objects.filter(created_at__gte=hace_7_dias).count()
        context['expedientes_creados_30d'] = Expediente.objects.filter(created_at__gte=hace_30_dias).count()

        usuarios = User.objects.all().select_related('profile').annotate(
            total_movimientos=Count('movimiento'),
            total_expedientes_asignados=Count('expediente'),
        ).order_by('-last_login')

        usuarios_data = []
        for u in usuarios:
            rol = u.profile.rol if hasattr(u, 'profile') else 'sin_perfil'
            ultimo_acceso = u.last_login
            hace_cuanto = ''
            if ultimo_acceso:
                diff = hoy - ultimo_acceso
                if diff.days == 0:
                    hace_cuanto = 'Hoy'
                elif diff.days == 1:
                    hace_cuanto = 'Ayer'
                else:
                    hace_cuanto = f'Hace {diff.days} días'

            usuarios_data.append({
                'user': u,
                'rol': rol,
                'ultimo_acceso': ultimo_acceso,
                'hace_cuanto': hace_cuanto,
                'total_movimientos': u.total_movimientos,
                'total_expedientes': u.total_expedientes_asignados,
                'es_staff': u.is_staff,
                'es_activo': u.is_active,
            })

        context['usuarios'] = usuarios_data

        context['actividad_reciente'] = Movimiento.objects.select_related(
            'expediente', 'usuario', 'usuario__profile'
        ).order_by('-created_at')[:30]

        montos = Expediente.objects.aggregate(
            total_reclamado=Sum('monto_reclamado'),
            total_convenio=Sum('monto_convenio'),
        )
        context['total_reclamado'] = montos['total_reclamado'] or 0
        context['total_convenio'] = montos['total_convenio'] or 0

        context['casos_por_estado'] = {
            label: Expediente.objects.filter(estado=key).count()
            for key, label in Expediente.ESTADO_CHOICES
        }

        context['fecha_generacion'] = hoy
        context['superadmin'] = self.request.user

        return context
